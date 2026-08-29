#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SEC Nigeria Weekly NAV for CIS extractor — V6 FIABLE COMPLETE
====================================================

Identifiant :
    SEC_NG_NAV_EXTRACTOR_V6_FULL_XLS_CONVERSION_RELIABLE_DATE_BLOCKS

Objectif :
    Télécharger les fichiers Excel hebdomadaires de la SEC Nigeria, extraire les fonds,
    les sociétés de gestion, les actifs nets / NAV, les prix de VL / Offer Price / Unit Price,
    les catégories et les devises, avec une attention particulière à la fiabilité de
    l'association DATE -> BLOC DE COLONNES -> DONNEES.

Améliorations majeures V4 par rapport à V3 :
    1. Correction renforcée de l'association des dates aux blocs de colonnes.
       Le script ne doit plus rattacher automatiquement les deux blocs d'un fichier
       à la date finale du fichier.

    2. Détection explicite des dates par "bandes" de colonnes.
       Les cellules fusionnées au-dessus d'un bloc de colonnes sont propagées.
       Chaque date visible au-dessus d'un bloc devient une candidate de rattachement.

    3. Règle robuste pour fichiers avec deux dates :
       Si deux blocs NAV existent et deux dates différentes sont visibles dans la feuille,
       le bloc de gauche reçoit la date de gauche et le bloc de droite reçoit la date de droite.

    4. Les dates issues du titre du fichier ou du lien SEC ne sont utilisées qu'en fallback.
       Une date dans la feuille est prioritaire sur une date dans le titre du fichier.

    5. Catégories simplifiées en français majuscule sans accent :
       ACTIONS, MONETAIRE, OBLIGATAIRE, DOLLAR, IMMOBILIER, DIVERSIFIE, ETHIQUE,
       CHARIA, SPECIALISE, INFRASTRUCTURE, ETF, AUTRE, NON CLASSE.

    6. Ajout d'un rapport de couverture annuelle :
       Une année complète doit normalement présenter environ 48 à 54 dates de valorisation
       distinctes. Le rapport signale les années incomplètes ou suspectes.

    7. Ajout de clés de dédoublonnage plus propres :
       valuation_date + fund_manager_key + fund_name_key + category + currency.

    8. Audit plus précis :
       dates détectées, dates extraites, nombre de dates par année, nombre de fichiers,
       nombre de lignes, conflits de NAV/VL.

Installation :
    pip install requests beautifulsoup4 openpyxl python-dateutil

Exemple — extraction 2026, 2025, 2024, 2023 :
    python sec_ng_nav_extractor_v4_reliable.py \
      --years 2026 2025 2024 2023 \
      --cache-dir sec_ng_downloads \
      --out sec_ng_nav_2023_2026_v4.csv \
      --audit sec_ng_nav_audit_2023_2026_v4.csv \
      --coherence sec_ng_nav_coherence_2023_2026_v4.csv \
      --coverage sec_ng_nav_annual_coverage_2023_2026_v4.csv \
      --fuzzy-report sec_ng_nav_fuzzy_names_2023_2026_v4.csv \
      --strict-quality

Exemple — extraction locale :
    python sec_ng_nav_extractor_v4_reliable.py \
      --local-files "NAV-as-at-10th-February-2023.xlsx" "NAV-as-at-10th-May-2024 (1).xlsx" \
      --out local_nav_v4.csv \
      --audit local_audit_v4.csv \
      --coherence local_coherence_v4.csv \
      --coverage local_coverage_v4.csv \
      --strict-quality

Sorties :
    - CSV principal : observations fonds/date/prix.
    - CSV audit : audit fichiers et feuilles.
    - CSV coherence : conflits ou confirmations entre fichiers.
    - CSV coverage : contrôle du nombre de dates par année.
    - CSV fuzzy-report : noms proches à revoir pour référentiel maître.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
import shutil
import subprocess
import tempfile
import time
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, asdict, field
from datetime import date, datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from dateutil import parser as dtparser
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# 0. PARAMETRES
# =============================================================================

BASE_WEEKLY_URL = (
    "https://sec.gov.ng/for-operators/keep-track-of-capital-market-data/"
    "net-asset-value-data/weekly-net-asset-value-for-cis/"
)

YEAR_URL_TEMPLATE = BASE_WEEKLY_URL + "{year}-weekly-nav-for-cis/"

DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; SEC-Nigeria-NAV-Extractor/4.0; "
        "+https://sec.gov.ng/)"
    )
}

MONTHS_PATTERN = (
    r"January|February|March|April|May|June|July|August|September|October|November|December|"
    r"Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec"
)

DEFAULT_PRICE_TOLERANCE = 0.0001
DEFAULT_NAV_TOLERANCE = 1.0

STOP_ROW_LABELS = {
    "SUB-TOTAL",
    "SUBTOTAL",
    "TOTAL",
    "GRAND TOTAL",
    "MOVING AVERAGE",
    "AVERAGE",
}

# Catégories de sortie : français, majuscule, sans accent, sans préfixe "FONDS".
CATEGORY_FR_RULES: List[Tuple[str, str]] = [
    ("MONEY MARKET", "MONETAIRE"),
    ("MONEY MKT", "MONETAIRE"),
    ("MMF", "MONETAIRE"),
    ("EQUITY", "ACTIONS"),
    ("STOCK", "ACTIONS"),
    ("SHARE", "ACTIONS"),
    ("BOND", "OBLIGATAIRE"),
    ("FIXED INCOME", "OBLIGATAIRE"),
    ("DOLLAR", "DOLLAR"),
    ("USD", "DOLLAR"),
    ("REAL ESTATE", "IMMOBILIER"),
    ("REIT", "IMMOBILIER"),
    ("BALANCED", "DIVERSIFIE"),
    ("MIXED", "DIVERSIFIE"),
    ("ETHICAL", "ETHIQUE"),
    ("SHARI", "CHARIA"),
    ("SUKUK", "CHARIA"),
    ("SPECIALISED", "SPECIALISE"),
    ("SPECIALIZED", "SPECIALISE"),
    ("INFRASTRUCTURE", "INFRASTRUCTURE"),
    ("ETF", "ETF"),
    ("EXCHANGE TRADED", "ETF"),
]

CATEGORY_KEYWORDS = tuple(k for k, _ in CATEGORY_FR_RULES)

CURRENCY_RULES: List[Tuple[str, str, str]] = [
    ("₦", "NGN", "NAIRA NIGERIAN"),
    (" NGN", "NGN", "NAIRA NIGERIAN"),
    ("NAIRA", "NGN", "NAIRA NIGERIAN"),
    ("NIGERIAN NAIRA", "NGN", "NAIRA NIGERIAN"),
    ("US$", "USD", "DOLLAR AMERICAIN"),
    ("USD", "USD", "DOLLAR AMERICAIN"),
    ("DOLLAR", "USD", "DOLLAR AMERICAIN"),
    ("EURO", "EUR", "EURO"),
    ("EUR", "EUR", "EURO"),
    ("GBP", "GBP", "LIVRE STERLING"),
    ("POUND", "GBP", "LIVRE STERLING"),
    # Place en DERNIER, apres « US$ » : les en-tetes SEC ecrivent parfois la
    # devise sous forme de symbole seul (« NAV ($) », « Offer Price ($) »).
    # Sans cette regle, ces colonnes retombaient sur l inference par le nom du
    # fonds. Le naira s ecrivant ₦ ou N dans ces memes fichiers, « $ » est
    # univoque ici — contrairement a « N », volontairement absent de cette liste
    # car bien trop generique.
    ("$", "USD", "DOLLAR AMERICAIN"),
]


# =============================================================================
# 1. STRUCTURES
# =============================================================================

@dataclass
class DateCandidate:
    sheet_name: str
    row_idx: int
    col_idx: int
    value: str
    parsed_date: str
    source: str
    confidence: int


@dataclass
class DateBand:
    """
    Bande de colonnes couverte par une date détectée dans la feuille.
    Exemple :
        date 2024-05-03 couvre les colonnes K:O
        date 2024-05-10 couvre les colonnes P:T
    """
    sheet_name: str
    parsed_date: str
    value: str
    source: str
    confidence: int
    row_idx: int
    start_col: int
    end_col: int
    center_col: float


@dataclass
class ColumnBlock:
    sheet_name: str
    block_id: int
    block_type: str

    valuation_date: str
    valuation_date_text: str
    valuation_date_source: str
    valuation_date_confidence: int
    valuation_date_row: int
    valuation_date_col: int
    date_assignment_method: str

    start_col: int
    end_col: int
    nav_col: Optional[int]
    pct_to_total_col: Optional[int]
    bid_price_col: Optional[int]
    offer_price_col: Optional[int]
    unit_price_col: Optional[int]
    unitholders_col: Optional[int]
    yield_wtd_col: Optional[int]
    yield_ytd_col: Optional[int]

    block_currency_code: str
    block_currency_name_fr: str
    block_currency_source: str
    block_currency_confidence: int

    # En-tetes bruts des colonnes de prix. La SEC publie les paires
    # « Offer Price (NGN) » / « Offer Price (USD) » : la devise d une mesure est
    # donc portee par l en-tete de la colonne d ou elle vient, et nulle part
    # ailleurs. Sans cette information la devise etait deduite du NOM DU FONDS,
    # d ou 238 lignes en naira etiquetees USD (mesure du lot AE).
    bid_price_header: str = ""
    offer_price_header: str = ""
    unit_price_header: str = ""

    # Toutes les colonnes de prix du bloc, sous la forme (type, index, devise).
    # Un bloc porte typiquement SIX colonnes de prix — Bid, Offer et Unit, en
    # dollar et en naira. N en retenir qu une par type revenait a choisir la
    # devise au hasard de l ordre des colonnes : c est l origine de #73.
    price_columns: List[Tuple[str, int, str]] = field(default_factory=list)


@dataclass
class NavRecord:
    # Source
    source_file: str
    source_url: str
    source_title: str
    source_page_url: str
    source_year_page: str
    downloaded_at_utc: str
    file_size_bytes: int

    # Feuille / emplacement
    sheet_name: str
    source_row_number: int
    header_row_number: int
    block_id: int
    block_type: str

    # Date
    valuation_date: str
    valuation_date_text: str
    valuation_date_source: str
    valuation_date_confidence: int
    valuation_date_row: int
    valuation_date_col: int
    date_assignment_method: str
    previous_or_current_hint: str
    year: str
    month: str
    iso_week: str

    # Fonds
    fund_name_raw: str
    fund_name_clean: str
    fund_name_key: str

    # Société de gestion
    fund_manager_raw: str
    fund_manager_clean: str
    fund_manager_key: str

    # Catégorie
    fund_category_raw: str
    fund_category_key: str
    fund_category_fr: str
    fund_category_confidence: int

    # Devise / monnaie
    currency_code: str
    currency_name_fr: str
    currency_source: str
    currency_confidence: int

    # Valeurs numériques
    nav_ngn: Any
    nav_value: Any
    nav_currency_code: str
    pct_to_total: Any
    bid_price: Any
    offer_price: Any
    unit_price: Any
    vl_price: Any
    vl_price_source: str
    vl_currency_code: str
    # D ou vient la devise de la VL : « column_header » quand elle est lue dans
    # l en-tete de la colonne utilisee (fiable), « inferred_* » quand elle est
    # seulement deduite du contexte (nom du fonds, categorie) — auquel cas elle
    # ne doit pas etre traitee comme une preuve en aval.
    vl_currency_source: str
    vl_currency_confidence: int
    unitholders: Any
    yield_wtd: Any
    yield_ytd: Any

    # Qualité
    observation_key: str
    quality_flags: str
    extraction_status: str

    # Prix par devise, INDEPENDAMMENT de la colonne retenue dans `vl_price`.
    # Emis pour que l aval puisse corriger vers une devise choisie sans jamais
    # convertir : une valeur lue, ou rien. Vide quand la SEC ne publie pas cette
    # devise ce jour-la — l absence est une information, pas un trou a combler.
    # Defaut vide : ces champs viennent APRES des champs sans defaut, ce qui
    # n est licite que dans cet ordre. Les placer avant leverait
    # « non-default argument follows default argument » — erreur qu `ast.parse`
    # ne detecte pas, comme le lot AI l a appris a ses depens.
    vl_price_ngn: Any = ""
    vl_price_ngn_source: str = ""
    vl_price_usd: Any = ""
    vl_price_usd_source: str = ""


@dataclass
class AuditRecord:
    source_file: str
    source_url: str
    source_title: str
    source_page_url: str
    source_year_page: str
    status: str
    message: str
    sheet_name: str = ""
    header_row: str = ""
    detected_date_count: int = 0
    detected_dates: str = ""
    detected_date_bands: str = ""
    detected_block_count: int = 0
    extracted_rows: int = 0
    extracted_dates: str = ""
    date_min: str = ""
    date_max: str = ""


@dataclass
class CoherenceRecord:
    coherence_key: str
    valuation_date: str
    fund_name_key: str
    fund_manager_key: str
    fund_category_fr: str
    currency_code: str
    occurrence_count: int
    source_files: str
    source_urls: str
    status: str
    nav_values: str
    vl_price_values: str
    message: str


@dataclass
class CoverageRecord:
    year: str
    distinct_valuation_dates: int
    first_date: str
    last_date: str
    expected_min_dates: int
    expected_max_dates: int
    status: str
    message: str
    dates: str
    source_files_count: int
    records_count: int


# =============================================================================
# 2. UTILITAIRES
# =============================================================================

def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\xa0", " ").replace("\u200b", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def strip_accents(text: str) -> str:
    decomposed = unicodedata.normalize("NFKD", text or "")
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def upper_no_accent(text: str) -> str:
    text = strip_accents(clean_text(text)).upper()
    text = re.sub(r"\s+", " ", text).strip()
    return text


def text_upper(value: Any) -> str:
    return upper_no_accent(clean_text(value))


def normalize_name_for_key(text: str) -> str:
    text = upper_no_accent(text)
    text = text.replace("&", " AND ")
    text = re.sub(r"['’`]", "", text)
    text = re.sub(r"[^A-Z0-9]+", " ", text)

    replacements = {
        r"\bLIMITED\b": "LTD",
        r"\bL T D\b": "LTD",
        r"\bPUBLIC LIMITED COMPANY\b": "PLC",
        r"\bP L C\b": "PLC",
        r"\bMONEY MKT\b": "MONEY MARKET",
        r"\bMMF\b": "MONEY MARKET FUND",
    }
    for pat, repl in replacements.items():
        text = re.sub(pat, repl, text)

    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_display_name(text: str) -> str:
    return re.sub(r"\s+", " ", clean_text(text)).strip()


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def safe_float(value: Any) -> Optional[float]:
    if is_number(value):
        return float(value)
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(",", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    try:
        return float(text)
    except Exception:
        return None


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_url(href: str, base_url: str) -> str:
    return urljoin(base_url, href)


def safe_filename_from_url(url: str, fallback: str = "download.xlsx") -> str:
    parsed = urlparse(url)
    name = Path(parsed.path).name or fallback
    name = re.sub(r"[^\w.\-() ]+", "_", name)
    if not name.lower().endswith((".xlsx", ".xlsm", ".xls")):
        name += ".xlsx"
    return name


def extract_dates_from_text(text: str) -> List[Tuple[str, str]]:
    original = clean_text(text)
    if not original:
        return []

    txt = original.replace("-", " ")
    txt = re.sub(r"(\d{1,2})(st|nd|rd|th)", r"\1", txt, flags=re.I)

    patterns = [
        rf"\b\d{{1,2}}\s+(?:{MONTHS_PATTERN})\s+\d{{4}}\b",
        rf"\b(?:{MONTHS_PATTERN})\s+\d{{1,2}},?\s+\d{{4}}\b",
    ]

    results: List[Tuple[str, str]] = []
    for pat in patterns:
        for m in re.finditer(pat, txt, flags=re.I):
            raw = m.group(0)
            parsed: Optional[date] = None
            for dayfirst in (True, False):
                try:
                    parsed = dtparser.parse(raw, dayfirst=dayfirst, fuzzy=True).date()
                    break
                except Exception:
                    parsed = None
            if parsed:
                results.append((raw, parsed.isoformat()))

    seen = set()
    unique: List[Tuple[str, str]] = []
    for raw, iso in results:
        key = (raw.upper(), iso)
        if key not in seen:
            unique.append((raw, iso))
            seen.add(key)

    return unique


def infer_previous_current_from_text(text: str) -> str:
    t = text_upper(text)
    if any(x in t for x in ("PREVIOUS", "PRIOR", "LAST WEEK", "LAST NAV", "PRECEDING")):
        return "PREVIOUS"
    if any(x in t for x in ("CURRENT", "THIS WEEK", "AS AT", "AS OF")):
        return "CURRENT"
    return ""


def date_parts(iso_date: str) -> Tuple[str, str, str]:
    if not iso_date:
        return "", "", ""
    try:
        d = datetime.strptime(iso_date, "%Y-%m-%d").date()
        return str(d.year), f"{d.month:02d}", f"{d.isocalendar().week:02d}"
    except Exception:
        return "", "", ""


# =============================================================================
# 3. CATEGORIES ET DEVISES
# =============================================================================

def classify_category_fr(raw_category: str, fund_name: str = "") -> Tuple[str, int]:
    blob_category = upper_no_accent(raw_category)
    blob_fund = upper_no_accent(fund_name)
    blob = f"{blob_category} {blob_fund}".strip()

    if not blob:
        return "NON CLASSE", 0

    for keyword, fr in CATEGORY_FR_RULES:
        if keyword in blob:
            confidence = 100 if keyword in blob_category else 60
            return upper_no_accent(fr), confidence

    return "AUTRE", 30


def is_category_row(row: Sequence[Any]) -> bool:
    values = [clean_text(x) for x in row if clean_text(x)]
    if not values:
        return False

    joined = upper_no_accent(" ".join(values))
    has_keyword = any(k in joined for k in CATEGORY_KEYWORDS)
    numeric_count = sum(1 for x in row if is_number(x))

    return has_keyword and numeric_count <= 1 and len(values) <= 5


def category_from_row(row: Sequence[Any]) -> str:
    for x in row:
        t = clean_text(x)
        if t:
            return t
    return ""


def detect_currency_from_text(text: str) -> Tuple[str, str, int]:
    if not text:
        return "", "", 0

    # La ponctuation doit devenir des separateurs AVANT la recherche.
    # Les marqueurs sont compares entoures d espaces (" USD "), si bien qu un
    # en-tete « Offer Price (USD) » ne correspondait a rien : les parentheses
    # collaient au code devise. Or c est exactement le format des fichiers SEC
    # (« Offer Price (NGN) » / « Offer Price (USD) »), donc la devise portee par
    # l en-tete de colonne etait systematiquement manquee — et l extraction
    # retombait sur l inference par le nom du fonds, a l origine de #73.
    # `$` et `₦` sont conserves : ils font partie des marqueurs (« US$ »).
    _texte = upper_no_accent(text).replace("₦", " ₦ ")
    blob = " " + re.sub(r"[^A-Z0-9$₦]+", " ", _texte).strip() + " "

    for marker, code, name_fr in CURRENCY_RULES:
        if marker == "₦":
            if "₦" in text:
                return code, name_fr, 100
        else:
            marker_norm = " " + upper_no_accent(marker).strip() + " "
            if marker_norm in blob:
                return code, name_fr, 95

    return "", "", 0


def detect_currency_in_column_header(header: str) -> str:
    """Devise portee par l en-tete d une colonne de prix.

    Les fichiers SEC ecrivent « Offer Price ($) » et « Offer Price (N) » : la
    devise tient dans un seul caractere entre parentheses, et les deux devises
    occupent des colonnes SEPAREES (mesure du lot AH sur le fichier du
    2026-07-24 : Bid Price ($) = 119,92 et Bid Price (N) = 165 509,54 pour le
    meme fonds).

    « N » est trop generique pour figurer dans CURRENCY_RULES, ou il
    provoquerait des faux positifs sur n importe quel texte. Entre parentheses
    et dans un en-tete de prix, il est en revanche univoque — d ou ce detecteur
    dedie, applique aux seuls en-tetes de colonnes.
    """
    h = upper_no_accent(header or "")
    if re.search(r"\(\s*(?:N|NGN|NAIRA)\s*\)", h):
        return "NGN"
    if re.search(r"\(\s*(?:\$|USD|US\$)\s*\)", h):
        return "USD"
    code, _, _ = detect_currency_from_text(header or "")
    return code


def infer_currency(
    category_raw: str,
    fund_name: str,
    fund_manager: str,
    header_text: str,
    block_text: str,
) -> Tuple[str, str, str, int]:
    checks = [
        ("header_or_block", f"{header_text} {block_text}", 100),
        ("category", category_raw, 90),
        ("fund_name", fund_name, 75),
        ("fund_manager", fund_manager, 40),
    ]

    for source, text, base_conf in checks:
        code, name, conf = detect_currency_from_text(text)
        if code:
            return code, name, source, min(conf, base_conf)

    cat_key = upper_no_accent(category_raw)
    fund_key = upper_no_accent(fund_name)
    if "DOLLAR" in cat_key or "USD" in cat_key or "DOLLAR" in fund_key or "USD" in fund_key:
        return "USD", "DOLLAR AMERICAIN", "category_or_fund_name", 80

    return "NGN", "NAIRA NIGERIAN", "default_nigeria_context", 50


# =============================================================================
# 4. TELECHARGEMENT
# =============================================================================

def fetch_year_links(year: int, timeout: int = 30) -> List[Dict[str, str]]:
    page_url = YEAR_URL_TEMPLATE.format(year=year)
    response = requests.get(page_url, headers=DEFAULT_HEADERS, timeout=timeout)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    links: List[Dict[str, str]] = []

    for a in soup.find_all("a", href=True):
        href = a.get("href", "")
        title = clean_text(a.get_text(" ", strip=True))
        full_url = normalize_url(href, page_url)
        blob = f"{href} {title}".lower()

        if (
            ".xlsx" in blob
            or ".xlsm" in blob
            or ".xls" in blob
            or ("/documents/" in blob and "asset value" in blob)
            or ("/documents/" in blob and "unit price" in blob)
            or ("/documents/" in blob and "nav" in blob)
        ):
            if "weekly-nav-for-cis" in full_url.lower():
                continue

            links.append(
                {
                    "year": str(year),
                    "title": title,
                    "url": full_url,
                    "page_url": page_url,
                    "year_page": page_url,
                }
            )

    dedup: Dict[str, Dict[str, str]] = {}
    for item in links:
        dedup[item["url"]] = item

    return list(dedup.values())


def download_file(url: str, cache_dir: Path, timeout: int = 60, sleep_seconds: float = 0.2) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    path = cache_dir / safe_filename_from_url(url)

    if path.exists() and path.stat().st_size > 1000:
        return path

    response = requests.get(url, headers=DEFAULT_HEADERS, timeout=timeout)
    response.raise_for_status()

    path.write_bytes(response.content)

    if sleep_seconds > 0:
        time.sleep(sleep_seconds)

    return path



# =============================================================================
# 4B. COMPATIBILITE ANCIENS FICHIERS .XLS
# =============================================================================

def is_legacy_xls(path: Path) -> bool:
    """
    Détecte les anciens fichiers Excel .xls.
    openpyxl ne sait pas lire directement les .xls binaires historiques.
    Pour les années anciennes SEC Nigeria, notamment 2018-2020, cette conversion
    est indispensable.
    """
    return path.suffix.lower() == ".xls"


def find_libreoffice_binary() -> Optional[str]:
    """
    Recherche LibreOffice / soffice sur la machine.
    La conversion .xls -> .xlsx s'appuie sur LibreOffice en mode headless.
    """
    for binary in ("libreoffice", "soffice"):
        found = shutil.which(binary)
        if found:
            return found
    return None


def convert_xls_to_xlsx_with_libreoffice(path: Path, converted_dir: Optional[Path] = None) -> Path:
    """
    Convertit un fichier .xls en .xlsx avec LibreOffice headless.

    Important :
    - ne modifie pas le fichier source ;
    - crée un fichier converti dans un dossier temporaire ou dans converted_dir ;
    - retourne le chemin du .xlsx converti ;
    - lève une erreur explicite si LibreOffice n'est pas disponible.

    Installation Linux/Debian/Ubuntu :
        sudo apt-get update
        sudo apt-get install -y libreoffice

    Installation serveur :
        libreoffice --headless --convert-to xlsx --outdir <dossier> fichier.xls
    """
    lo = find_libreoffice_binary()
    if not lo:
        raise RuntimeError(
            "Fichier .xls détecté mais LibreOffice/soffice est introuvable. "
            "Installer LibreOffice ou convertir les .xls en .xlsx avant exécution."
        )

    if converted_dir is None:
        converted_dir = path.parent / "_converted_xlsx"
    converted_dir.mkdir(parents=True, exist_ok=True)

    expected = converted_dir / (path.stem + ".xlsx")
    if expected.exists() and expected.stat().st_size > 1000:
        return expected

    cmd = [
        lo,
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(converted_dir),
        str(path),
    ]

    proc = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=120,
    )

    if proc.returncode != 0:
        raise RuntimeError(
            "Echec de conversion .xls -> .xlsx avec LibreOffice. "
            f"Commande={cmd} STDOUT={proc.stdout} STDERR={proc.stderr}"
        )

    if not expected.exists():
        # LibreOffice peut parfois produire un nom légèrement différent.
        candidates = sorted(converted_dir.glob(path.stem + "*.xlsx"))
        if candidates:
            expected = candidates[0]

    if not expected.exists() or expected.stat().st_size <= 1000:
        raise RuntimeError(
            "Conversion .xls -> .xlsx terminee mais fichier converti introuvable ou vide. "
            f"Sortie attendue={expected} STDOUT={proc.stdout} STDERR={proc.stderr}"
        )

    return expected


def prepare_workbook_for_openpyxl(path: Path) -> Tuple[Path, str]:
    """
    Retourne un fichier lisible par openpyxl et un statut de préparation.
    - .xlsx/.xlsm : fichier original ;
    - .xls : conversion automatique en .xlsx via LibreOffice.
    """
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return path, "ORIGINAL_OPENPYXL_COMPATIBLE"

    if suffix == ".xls":
        converted = convert_xls_to_xlsx_with_libreoffice(path)
        return converted, f"CONVERTED_XLS_TO_XLSX:{converted.name}"

    raise RuntimeError(f"Format Excel non supporté : {path.suffix}")


# =============================================================================
# 5. EXCEL
# =============================================================================

def expanded_sheet_matrix(ws: Worksheet) -> List[List[Any]]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    matrix = [
        [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]

    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col_m, max_row_m = merged.bounds
        value = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                matrix[r - 1][c - 1] = value

    return matrix


def norm_header(value: Any) -> str:
    txt = text_upper(value)
    txt = txt.replace("₦", "N")
    txt = txt.replace("  ", " ")
    return txt


def detect_header_row(matrix: List[List[Any]], max_scan_rows: int = 100) -> Optional[int]:
    best_idx = None
    best_score = -1

    for idx, row in enumerate(matrix[:max_scan_rows]):
        headers = [norm_header(x) for x in row]
        joined = " | ".join(h for h in headers if h)

        score = 0

        if any(h in ("FUND", "FUND NAME", "NAME OF FUND") for h in headers):
            score += 8
        if any("FUND MANAGER" in h or h in ("MANAGER", "ASSET MANAGER") for h in headers):
            score += 8
        if any(h == "NAV" or "NET ASSET VALUE" in h or re.search(r"\bNAV\b", h) for h in headers):
            score += 8
        if any("OFFER" in h or "UNIT PRICE" in h or "BID" in h for h in headers):
            score += 6
        if any("UNITHOLDER" in h or "UNIHOLDER" in h for h in headers):
            score += 3
        if any("%" in h and "TOTAL" in h for h in headers):
            score += 2
        if "S/N" in headers or "SN" in headers or "SNO" in headers:
            score += 1

        if "AS AT" in joined and score < 16:
            score -= 5

        if score > best_score:
            best_score = score
            best_idx = idx

    if best_score >= 14:
        return best_idx

    return None


def score_sheet_for_data(name: str, matrix: List[List[Any]]) -> Tuple[int, Optional[int]]:
    header_idx = detect_header_row(matrix)
    lname = name.lower()

    score = 0

    if header_idx is not None:
        score += 100
    if "weekly valuation" in lname:
        score += 60
    if lname.strip() == "data":
        score += 55
    if "valuation" in lname:
        score += 40
    if "volatility" in lname:
        score += 30
    if "trend" in lname or "market share" in lname or "comparison" in lname:
        score -= 25

    if header_idx is not None:
        numeric_rows = 0
        for row in matrix[header_idx + 1 : header_idx + 100]:
            if sum(1 for x in row if is_number(x)) >= 2 and any(clean_text(x) for x in row):
                numeric_rows += 1
        score += min(numeric_rows, 40)

    return score, header_idx


def choose_candidate_sheets(wb, include_all_candidate_sheets: bool = False) -> List[Tuple[str, List[List[Any]], int]]:
    candidates: List[Tuple[int, str, List[List[Any]], int]] = []

    for ws in wb.worksheets:
        matrix = expanded_sheet_matrix(ws)
        score, header_idx = score_sheet_for_data(ws.title, matrix)
        if header_idx is not None and score >= 90:
            candidates.append((score, ws.title, matrix, header_idx))

    candidates.sort(key=lambda x: x[0], reverse=True)

    if not candidates:
        raise RuntimeError("Aucune feuille de données exploitable détectée.")

    if include_all_candidate_sheets:
        return [(name, matrix, header_idx) for _, name, matrix, header_idx in candidates]

    best = candidates[0]
    return [(best[1], best[2], best[3])]


# =============================================================================
# 6. DATES ET BANDES DE DATES
# =============================================================================

def extract_sheet_date_candidates(
    sheet_name: str,
    matrix: List[List[Any]],
    header_idx: int,
    source_title: str,
    source_file: str,
    source_url: str,
    top_rows_padding: int = 4,
) -> List[DateCandidate]:
    candidates: List[DateCandidate] = []
    max_row_to_scan = min(len(matrix), max(header_idx + 1 + top_rows_padding, 12))

    for r_idx in range(0, max_row_to_scan):
        row = matrix[r_idx]
        for c_idx, value in enumerate(row):
            text = clean_text(value)
            if not text:
                continue

            for raw, iso in extract_dates_from_text(text):
                confidence = 100
                if r_idx <= header_idx:
                    confidence += 20
                if "AS AT" in text_upper(text) or "AS OF" in text_upper(text):
                    confidence += 15
                if "CURRENT" in text_upper(text):
                    confidence += 10
                if "PREVIOUS" in text_upper(text):
                    confidence += 5

                candidates.append(
                    DateCandidate(
                        sheet_name=sheet_name,
                        row_idx=r_idx,
                        col_idx=c_idx,
                        value=text,
                        parsed_date=iso,
                        source="sheet_cell",
                        confidence=confidence,
                    )
                )

    for source_name, source_text, conf in [
        ("source_title", source_title, 80),
        ("source_file", source_file, 75),
        ("source_url", source_url, 70),
    ]:
        for raw, iso in extract_dates_from_text(source_text):
            candidates.append(
                DateCandidate(
                    sheet_name=sheet_name,
                    row_idx=-1,
                    col_idx=-1,
                    value=raw,
                    parsed_date=iso,
                    source=source_name,
                    confidence=conf,
                )
            )

    seen = set()
    out: List[DateCandidate] = []

    for c in candidates:
        key = (c.sheet_name, c.row_idx, c.col_idx, c.parsed_date, c.source)
        if key not in seen:
            out.append(c)
            seen.add(key)

    return out


def build_date_bands(
    sheet_name: str,
    matrix: List[List[Any]],
    header_idx: int,
    date_candidates: List[DateCandidate],
    max_cols: int,
) -> List[DateBand]:
    """
    Transforme les dates détectées dans la feuille en bandes de colonnes.

    Logique :
        - On ne prend que les dates venant de sheet_cell.
        - On garde les dates situées au-dessus ou très proches de l'en-tête.
        - On groupe par (row_idx, parsed_date, value).
        - Comme les cellules fusionnées ont été propagées, une même date apparaît
          souvent sur plusieurs colonnes : on obtient donc naturellement start_col/end_col.
        - Si une date n'apparaît qu'une fois, la bande est ponctuelle mais reste utile.
    """
    sheet_candidates = [
        c for c in date_candidates
        if c.source == "sheet_cell" and 0 <= c.row_idx <= header_idx + 3
    ]

    groups: Dict[Tuple[int, str, str], List[DateCandidate]] = defaultdict(list)
    for c in sheet_candidates:
        groups[(c.row_idx, c.parsed_date, c.value)].append(c)

    bands: List[DateBand] = []
    for (row_idx, parsed_date, value), items in groups.items():
        cols = sorted({c.col_idx for c in items})
        start_col = min(cols)
        end_col = max(cols)
        center_col = (start_col + end_col) / 2
        confidence = max(c.confidence for c in items)
        bands.append(
            DateBand(
                sheet_name=sheet_name,
                parsed_date=parsed_date,
                value=value,
                source="sheet_cell",
                confidence=confidence,
                row_idx=row_idx,
                start_col=start_col,
                end_col=end_col,
                center_col=center_col,
            )
        )

    # Dédupliquer les bandes très similaires.
    dedup: List[DateBand] = []
    for b in sorted(bands, key=lambda x: (x.row_idx, x.start_col, x.end_col, x.parsed_date)):
        duplicate = False
        for d in dedup:
            if (
                d.parsed_date == b.parsed_date
                and d.row_idx == b.row_idx
                and abs(d.center_col - b.center_col) <= 1
            ):
                duplicate = True
                break
        if not duplicate:
            dedup.append(b)

    return dedup


def assign_date_to_blocks_by_bands(
    blocks_meta: List[Dict[str, Any]],
    date_bands: List[DateBand],
    date_candidates: List[DateCandidate],
    header_idx: int,
) -> List[Dict[str, Any]]:
    """
    Point critique V6.

    Cette fonction affecte les dates aux blocs de colonnes de manière fiable.

    Priorité :
        1. Bande de date qui couvre directement la colonne NAV du bloc.
        2. Bande de date dont le centre est le plus proche du centre du bloc.
        3. Si nombre de dates uniques dans la feuille >= nombre de blocs :
           mapping gauche -> droite.
        4. Date source_title/source_file/source_url seulement en dernier recours.

    Cela corrige le problème rencontré dans la version précédente :
        bloc précédent et bloc courant recevaient parfois la même date finale.
    """
    if not blocks_meta:
        return blocks_meta

    # Trier les blocs de gauche à droite.
    blocks_meta = sorted(blocks_meta, key=lambda b: (b["start_col"], b["end_col"]))

    # Garder les bandes exploitables.
    usable_bands = sorted(
        date_bands,
        key=lambda b: (b.start_col, b.end_col, b.row_idx, b.parsed_date)
    )

    unique_dates_left_to_right: List[DateBand] = []
    seen_dates = set()
    for b in usable_bands:
        if b.parsed_date not in seen_dates:
            unique_dates_left_to_right.append(b)
            seen_dates.add(b.parsed_date)

    assigned_dates = []

    # Cas idéal : au moins autant de dates visibles que de blocs.
    # On force un mapping gauche -> droite si cela est cohérent.
    if len(unique_dates_left_to_right) >= len(blocks_meta) and len(blocks_meta) >= 2:
        for idx, bm in enumerate(blocks_meta):
            band = unique_dates_left_to_right[idx]
            bm["date_band"] = band
            bm["date_assignment_method"] = "left_to_right_sheet_date_bands"
            assigned_dates.append(band.parsed_date)
        return blocks_meta

    # Sinon, rattachement par couverture ou proximité.
    for bm in blocks_meta:
        nav_col = bm.get("nav_col")
        block_center = (bm["start_col"] + bm["end_col"]) / 2

        scored: List[Tuple[float, DateBand, str]] = []
        for band in usable_bands:
            covers_nav = nav_col is not None and band.start_col <= nav_col <= band.end_col
            overlaps_block = not (band.end_col < bm["start_col"] or band.start_col > bm["end_col"])
            distance = abs(band.center_col - block_center)

            score = float(band.confidence)
            method = "nearest_sheet_date_band"

            if covers_nav:
                score += 1000
                method = "covered_nav_by_sheet_date_band"
            elif overlaps_block:
                score += 500
                method = "overlapped_block_by_sheet_date_band"
            else:
                score -= distance * 5

            # Date au-dessus de l'en-tête favorisée.
            if band.row_idx <= header_idx:
                score += 50

            scored.append((score, band, method))

        if scored:
            scored.sort(key=lambda x: x[0], reverse=True)
            best_score, best_band, method = scored[0]
            if best_score >= 50:
                bm["date_band"] = best_band
                bm["date_assignment_method"] = method
                assigned_dates.append(best_band.parsed_date)
                continue

        # Dernier fallback : source title/file/url.
        source_candidates = [c for c in date_candidates if c.source != "sheet_cell"]
        source_candidates.sort(key=lambda c: c.confidence, reverse=True)
        if source_candidates:
            c = source_candidates[0]
            bm["date_band"] = DateBand(
                sheet_name=c.sheet_name,
                parsed_date=c.parsed_date,
                value=c.value,
                source=c.source,
                confidence=c.confidence,
                row_idx=c.row_idx,
                start_col=c.col_idx,
                end_col=c.col_idx,
                center_col=float(c.col_idx),
            )
            bm["date_assignment_method"] = "fallback_source_title_file_url"
        else:
            bm["date_band"] = None
            bm["date_assignment_method"] = "no_date_found"

    return blocks_meta


# =============================================================================
# 7. COLONNES ET BLOCS
# =============================================================================

def find_column(
    headers: List[str],
    exact: Sequence[str] = (),
    contains_any: Sequence[str] = (),
    contains_all: Sequence[str] = (),
) -> Optional[int]:
    exact_set = {x.upper() for x in exact}

    for i, h in enumerate(headers):
        if h in exact_set:
            return i

    for i, h in enumerate(headers):
        if contains_all and all(term.upper() in h for term in contains_all):
            return i

    for i, h in enumerate(headers):
        if contains_any and any(term.upper() in h for term in contains_any):
            return i

    return None


def detect_identity_columns(header_row: Sequence[Any]) -> Dict[str, Optional[int]]:
    headers = [norm_header(x) for x in header_row]

    fund_col = find_column(headers, exact=("FUND", "FUND NAME", "NAME OF FUND"))
    manager_col = find_column(headers, contains_all=("FUND", "MANAGER"))

    if manager_col is None:
        manager_col = find_column(headers, exact=("MANAGER", "ASSET MANAGER"))

    return {
        "fund_col": fund_col,
        "manager_col": manager_col,
    }


def is_nav_header(h: str) -> bool:
    h = norm_header(h)
    return (
        h == "NAV"
        or h == "NET ASSET VALUE"
        or bool(re.search(r"\bNAV\b", h))
    ) and "CHANGE" not in h and "DIFFERENCE" not in h and "%" not in h


def is_offer_header(h: str) -> bool:
    h = norm_header(h)
    return "OFFER" in h and "PRICE" in h


def is_unit_price_header(h: str) -> bool:
    h = norm_header(h)
    return "UNIT" in h and "PRICE" in h


def is_bid_header(h: str) -> bool:
    h = norm_header(h)
    return "BID" in h and "PRICE" in h


def is_pct_total_header(h: str) -> bool:
    h = norm_header(h)
    return "%" in h and "TOTAL" in h


def is_unitholders_header(h: str) -> bool:
    h = norm_header(h)
    return "UNITHOLDER" in h or "UNIHOLDER" in h or "UNIT HOLDER" in h


def is_yield_wtd_header(h: str) -> bool:
    h = norm_header(h)
    return "YIELD" in h and ("WTD" in h or "WEEK" in h or "WYD" in h)


def is_yield_ytd_header(h: str) -> bool:
    h = norm_header(h)
    return "YIELD" in h and "YTD" in h


def header_context_text(matrix: List[List[Any]], header_idx: int, start_col: int, end_col: int) -> str:
    texts: List[str] = []
    top = max(0, header_idx - 5)
    bottom = min(len(matrix), header_idx + 2)

    for r in range(top, bottom):
        row = matrix[r]
        for c in range(max(0, start_col), min(len(row), end_col + 1)):
            t = clean_text(row[c])
            if t:
                texts.append(t)

    return " | ".join(texts)


def detect_standard_blocks(
    sheet_name: str,
    matrix: List[List[Any]],
    header_idx: int,
    date_candidates: List[DateCandidate],
    date_bands: List[DateBand],
) -> List[ColumnBlock]:
    headers_raw = matrix[header_idx]
    headers = [norm_header(x) for x in headers_raw]
    nav_cols = [i for i, h in enumerate(headers) if is_nav_header(h)]

    if not nav_cols:
        return []

    blocks_meta: List[Dict[str, Any]] = []

    for pos, nav_col in enumerate(nav_cols):
        next_nav = nav_cols[pos + 1] if pos + 1 < len(nav_cols) else len(headers)
        start = nav_col
        end = next_nav - 1

        pct_col = bid_col = offer_col = unit_col = unitholders_col = y_wtd_col = y_ytd_col = None
        # Toutes les colonnes de prix du bloc, avec la devise lue dans leur
        # en-tete. Les variables *_col ci-dessus ne retiennent que la PREMIERE
        # de chaque type et sont conservees pour compatibilite.
        prix_cols: List[Tuple[str, int, str]] = []

        for c in range(start, next_nav):
            h = headers[c]
            if is_unit_price_header(h):
                prix_cols.append(("unit_price", c, detect_currency_in_column_header(h)))
            elif is_offer_header(h):
                prix_cols.append(("offer_price", c, detect_currency_in_column_header(h)))
            elif is_bid_header(h):
                prix_cols.append(("bid_price", c, detect_currency_in_column_header(h)))
            if c != nav_col and is_pct_total_header(h) and pct_col is None:
                pct_col = c
            if is_bid_header(h) and bid_col is None:
                bid_col = c
            if is_offer_header(h) and offer_col is None:
                offer_col = c
            if is_unit_price_header(h) and unit_col is None:
                unit_col = c
            if is_unitholders_header(h) and unitholders_col is None:
                unitholders_col = c
            if is_yield_wtd_header(h) and y_wtd_col is None:
                y_wtd_col = c
            if is_yield_ytd_header(h) and y_ytd_col is None:
                y_ytd_col = c

        # V5 quality rule: ignore variation/difference blocks that may reuse the header
        # name NAV but contain percentage changes instead of valuation observations.
        block_context_upper = upper_no_accent(header_context_text(matrix, header_idx, start, end))
        headers_context_upper = upper_no_accent(" ".join(headers[start:end+1]))
        change_block_markers = (
            "% CHANGE",
            "CHANGE CURRENT FROM PREVIOUS",
            "CURRENT FROM PREVIOUS",
            "DIFFERENCE",
            "VARIATION",
        )
        if any(marker in block_context_upper or marker in headers_context_upper for marker in change_block_markers):
            continue

        # V5 quality rule: a proper block should normally contain at least one price column.
        # Otherwise it is likely an aggregate/change block and is excluded.
        if offer_col is None and unit_col is None and bid_col is None:
            continue

        blocks_meta.append(
            {
                "pos": len(blocks_meta),
                "start_col": start,
                "end_col": end,
                "nav_col": nav_col,
                "pct_to_total_col": pct_col,
                "bid_price_col": bid_col,
                "offer_price_col": offer_col,
                "unit_price_col": unit_col,
                "bid_price_header": headers[bid_col] if bid_col is not None else "",
                "offer_price_header": headers[offer_col] if offer_col is not None else "",
                "unit_price_header": headers[unit_col] if unit_col is not None else "",
                "price_columns": prix_cols,
                "unitholders_col": unitholders_col,
                "yield_wtd_col": y_wtd_col,
                "yield_ytd_col": y_ytd_col,
            }
        )

    blocks_meta = assign_date_to_blocks_by_bands(
        blocks_meta=blocks_meta,
        date_bands=date_bands,
        date_candidates=date_candidates,
        header_idx=header_idx,
    )

    blocks: List[ColumnBlock] = []
    for bm in blocks_meta:
        band: Optional[DateBand] = bm.get("date_band")
        start = bm["start_col"]
        end = bm["end_col"]

        context = header_context_text(matrix, header_idx, start, end)
        headers_context = " ".join(headers[start:end+1])
        hint = infer_previous_current_from_text(context)

        if not hint:
            # Deux blocs classiques : gauche = previous, droite = current.
            if len(blocks_meta) == 2:
                hint = "PREVIOUS" if bm["pos"] == 0 else "CURRENT"
            else:
                hint = "CURRENT" if bm["pos"] == len(blocks_meta) - 1 else "HISTORICAL"

        cur_code, cur_name, cur_source, cur_conf = infer_currency(
            category_raw="",
            fund_name="",
            fund_manager="",
            header_text=headers_context,
            block_text=context,
        )

        blocks.append(
            ColumnBlock(
                sheet_name=sheet_name,
                block_id=len(blocks) + 1,
                block_type=hint,
                valuation_date=band.parsed_date if band else "",
                valuation_date_text=band.value if band else "",
                valuation_date_source=band.source if band else "",
                valuation_date_confidence=band.confidence if band else 0,
                valuation_date_row=band.row_idx if band else -1,
                valuation_date_col=band.start_col if band else -1,
                date_assignment_method=bm.get("date_assignment_method", "unknown"),
                start_col=start,
                end_col=end,
                nav_col=bm["nav_col"],
                pct_to_total_col=bm["pct_to_total_col"],
                bid_price_col=bm["bid_price_col"],
                offer_price_col=bm["offer_price_col"],
                unit_price_col=bm["unit_price_col"],
                bid_price_header=bm.get("bid_price_header", ""),
                offer_price_header=bm.get("offer_price_header", ""),
                unit_price_header=bm.get("unit_price_header", ""),
                price_columns=bm.get("price_columns", []),
                unitholders_col=bm["unitholders_col"],
                yield_wtd_col=bm["yield_wtd_col"],
                yield_ytd_col=bm["yield_ytd_col"],
                block_currency_code=cur_code,
                block_currency_name_fr=cur_name,
                block_currency_source=cur_source,
                block_currency_confidence=cur_conf,
            )
        )

    return blocks


def detect_volatility_or_date_price_blocks(
    sheet_name: str,
    matrix: List[List[Any]],
    header_idx: int,
    date_candidates: List[DateCandidate],
    date_bands: List[DateBand],
) -> List[ColumnBlock]:
    identity = detect_identity_columns(matrix[header_idx])
    identity_cols = {c for c in identity.values() if c is not None}

    blocks: List[ColumnBlock] = []

    # Utilise directement les bandes de dates.
    for band in sorted(date_bands, key=lambda b: (b.start_col, b.end_col)):
        # Éviter les colonnes identitaires.
        if band.start_col in identity_cols and band.end_col in identity_cols:
            continue

        # Choisir la meilleure colonne numérique dans la bande.
        best_col = None
        best_count = -1

        for col_idx in range(band.start_col, band.end_col + 1):
            numeric_count = 0
            for row in matrix[header_idx + 1 : header_idx + 100]:
                if col_idx < len(row) and is_number(row[col_idx]):
                    numeric_count += 1
            if numeric_count > best_count:
                best_col = col_idx
                best_count = numeric_count

        if best_col is None or best_count < 5:
            continue

        context = header_context_text(matrix, header_idx, band.start_col, band.end_col)
        cur_code, cur_name, cur_source, cur_conf = infer_currency(
            category_raw="",
            fund_name="",
            fund_manager="",
            header_text=context,
            block_text=context,
        )

        blocks.append(
            ColumnBlock(
                sheet_name=sheet_name,
                block_id=len(blocks) + 1,
                block_type="DATE_PRICE_MATRIX",
                valuation_date=band.parsed_date,
                valuation_date_text=band.value,
                valuation_date_source=band.source,
                valuation_date_confidence=band.confidence,
                valuation_date_row=band.row_idx,
                valuation_date_col=band.start_col,
                date_assignment_method="volatility_date_band",
                start_col=band.start_col,
                end_col=band.end_col,
                nav_col=None,
                pct_to_total_col=None,
                bid_price_col=None,
                offer_price_col=None,
                unit_price_col=best_col,
                unitholders_col=None,
                yield_wtd_col=None,
                yield_ytd_col=None,
                block_currency_code=cur_code,
                block_currency_name_fr=cur_name,
                block_currency_source=cur_source,
                block_currency_confidence=cur_conf,
            )
        )

    return blocks


# =============================================================================
# 8. EXTRACTION LIGNES
# =============================================================================

def row_non_empty_values(row: Sequence[Any]) -> List[str]:
    return [clean_text(x) for x in row if clean_text(x)]


def get_cell(row: Sequence[Any], idx: Optional[int]) -> Any:
    if idx is None:
        return None
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def is_stop_label(text: str) -> bool:
    cleaned = normalize_name_for_key(text)
    for label in STOP_ROW_LABELS:
        if cleaned == normalize_name_for_key(label):
            return True
    return False


def choose_vl_price(offer_price: Any, unit_price: Any, bid_price: Any) -> Tuple[Any, str]:
    """Retient la mesure la plus proche d une valeur liquidative.

    Le prix unitaire EST la VL ; le Bid est un prix de rachat et l Offer un prix
    de souscription — ni l un ni l autre n est une VL. La priorite retenait
    pourtant `offer_price` en premier, si bien que 100 % des lignes mesurees au
    lot AE portaient `vl_price_source = offer_price`, y compris quand un prix
    unitaire explicite existait dans le meme bloc.

    La BIBLE Nigeria l interdit : « Ne choisis pas silencieusement Bid ou Offer
    comme VL. » On ne peut pas laisser la valeur nulle ici sans vider l export,
    mais on peut cesser de preferer l Offer au prix unitaire, et surtout nommer
    explicitement le repli pour que l aval puisse le refuser en connaissance de
    cause.
    """
    if safe_float(unit_price) is not None:
        return unit_price, "unit_price"
    if safe_float(offer_price) is not None:
        return offer_price, "offer_price_fallback"
    if safe_float(bid_price) is not None:
        return bid_price, "bid_price_fallback"
    return None, ""


def price_in_currency(
    valeurs: Dict[int, Any],
    price_columns: List[Tuple[str, int, str]],
    devise_voulue: str,
) -> Tuple[Any, str]:
    """Le prix publie dans UNE devise donnee, quelle que soit la colonne retenue.

    POURQUOI CETTE FONCTION EXISTE
    ------------------------------
    `choose_price_column` retient UNE colonne — celle dont la devise correspond
    au fonds — et l extracteur n emet que celle-la. Or corriger l historique
    demande le prix dans une devise PRECISE, choisie en aval et non par
    l extracteur.

    Mesure du 2026-08-29 : la base contient 233 ruptures d echelle, dont 208 ont
    une source SEC identifiable. Les corriger vers le naira supposait de disposer
    du prix naira de chaque semaine — introuvable dans le CSV, qui ne portait que
    la colonne retenue, parfois le dollar. Sans cette fonction, la seule maniere
    d obtenir un naira aurait ete de diviser un dollar par un taux : fabriquer
    une valeur que personne n a publiee. La regle du projet l interdit, et a
    juste titre.

    L extracteur connait pourtant deja toutes les colonnes et leurs devises : il
    ne les exposait simplement pas. Cette fonction ne decouvre rien, elle rend
    lisible ce qui etait deja lu.

    Meme ordre de priorite que `choose_price_column` — le prix unitaire EST la
    VL, Bid et Offer sont des replis — pour que les deux sorties soient
    comparables. Une priorite differente rendrait les deux colonnes du CSV
    incoherentes entre elles sans que rien ne le signale.

    Retourne (valeur, source_du_prix). (None, "") si cette devise n est pas
    publiee : une absence se dit, elle ne se comble pas.
    """
    if not devise_voulue:
        return None, ""
    rang = {"unit_price": 0, "offer_price": 1, "bid_price": 2}
    for kind, col, devise in sorted(price_columns, key=lambda x: rang.get(x[0], 9)):
        if devise != devise_voulue:
            continue
        v = valeurs.get(col)
        if safe_float(v) is not None:
            return v, kind if kind == "unit_price" else f"{kind}_fallback"
    return None, ""


def choose_price_column(
    valeurs: Dict[int, Any],
    price_columns: List[Tuple[str, int, str]],
    devise_du_fonds: str,
) -> Tuple[Any, str, str, str]:
    """Retient la mesure la plus juste parmi les colonnes de prix du bloc.

    Chaque signal est utilise pour ce qu il sait faire : le NOM du fonds dit sa
    devise de libelle, l EN-TETE dit la devise de chaque colonne. Il suffit donc
    de prendre la colonne dont la devise correspond au fonds.

    Mesure du lot AH, fichier du 2026-07-24, Afrinvest Dollar Fund :
        Bid Price ($) = 119,9184      Bid Price (N) = 165 509,54
    La source est propre et publie les deux. L extracteur retenait la colonne
    naira puis l etiquetait USD d apres le nom du fonds — d ou les series
    melangeant deux echelles de #73.

    Retourne (valeur, source_du_prix, devise, provenance_de_la_devise).
    """
    rang = {"unit_price": 0, "offer_price": 1, "bid_price": 2}
    candidats = sorted(price_columns, key=lambda x: rang.get(x[0], 9))

    def _nom_source(kind: str) -> str:
        # Un prix unitaire EST la VL ; Bid et Offer sont des replis, nommes
        # comme tels pour que l aval puisse les refuser sciemment.
        return kind if kind == "unit_price" else f"{kind}_fallback"

    # 1. La colonne dont la devise est celle du fonds.
    if devise_du_fonds:
        for kind, col, devise in candidats:
            if devise and devise == devise_du_fonds:
                v = valeurs.get(col)
                if safe_float(v) is not None:
                    return v, _nom_source(kind), devise, "column_header_matched_fund"

    # 2. A defaut, toute colonne exploitable — etiquetee par SA propre devise,
    #    jamais par celle du fonds.
    for kind, col, devise in candidats:
        v = valeurs.get(col)
        if safe_float(v) is not None:
            return v, _nom_source(kind), devise, "column_header" if devise else "unknown_column_currency"

    return None, "", "", ""


def make_observation_key(record_like: Dict[str, Any]) -> str:
    return "|".join(
        [
            clean_text(record_like.get("valuation_date")),
            clean_text(record_like.get("fund_manager_key")),
            clean_text(record_like.get("fund_name_key")),
            clean_text(record_like.get("fund_category_fr")),
            clean_text(record_like.get("currency_code")),
        ]
    )


def validate_record_quality(record: NavRecord) -> str:
    flags = []

    if not record.valuation_date:
        flags.append("MISSING_VALUATION_DATE")
    if safe_float(record.vl_price) is None:
        flags.append("MISSING_VL_PRICE")
    if not record.fund_name_key:
        flags.append("MISSING_FUND_NAME")
    if not record.fund_manager_key:
        flags.append("MISSING_FUND_MANAGER")
    if record.valuation_date_confidence < 80:
        flags.append("LOW_DATE_CONFIDENCE")
    if record.date_assignment_method.startswith("fallback"):
        flags.append("DATE_FROM_SOURCE_FALLBACK")
    if record.vl_price_source == "bid_price_fallback":
        flags.append("BID_PRICE_USED_AS_FALLBACK")
    if record.vl_price_source == "offer_price_fallback":
        # Un prix de souscription n est pas une VL. La ligne reste exportee,
        # mais l aval doit pouvoir la refuser en connaissance de cause.
        flags.append("OFFER_PRICE_USED_AS_FALLBACK")
    if record.vl_currency_source and record.vl_currency_source.startswith("inferred_"):
        # La devise n a pas pu etre lue dans l en-tete de colonne : elle est
        # deduite du contexte et ne constitue pas une preuve.
        flags.append("CURRENCY_INFERRED_NOT_FROM_COLUMN")
    if (
        record.vl_currency_code
        and record.currency_code
        and record.vl_currency_code != record.currency_code
    ):
        # L en-tete de colonne contredit l inference de contexte. C est
        # exactement le defaut mesure au lot AE : « Afrinvest Dollar Fund »
        # etait etiquete USD par son nom alors que la valeur venait d une
        # colonne en naira.
        flags.append("CURRENCY_COLUMN_DIFFERS_FROM_CONTEXT")
    if not record.fund_category_fr or record.fund_category_fr in {"NON CLASSE", "AUTRE"}:
        flags.append("CATEGORY_TO_REVIEW")
    if not record.currency_code:
        flags.append("MISSING_CURRENCY")
    if record.currency_confidence < 60:
        flags.append("LOW_CURRENCY_CONFIDENCE")

    return "|".join(flags)


def parse_sheet_records(
    path: Path,
    source_url: str,
    source_title: str,
    source_page_url: str,
    source_year_page: str,
    downloaded_at_utc: str,
    sheet_name: str,
    matrix: List[List[Any]],
    header_idx: int,
) -> Tuple[List[NavRecord], List[DateCandidate], List[DateBand], List[ColumnBlock]]:
    date_candidates = extract_sheet_date_candidates(
        sheet_name=sheet_name,
        matrix=matrix,
        header_idx=header_idx,
        source_title=source_title,
        source_file=path.name,
        source_url=source_url,
    )

    max_cols = max((len(r) for r in matrix), default=0)
    date_bands = build_date_bands(
        sheet_name=sheet_name,
        matrix=matrix,
        header_idx=header_idx,
        date_candidates=date_candidates,
        max_cols=max_cols,
    )

    identity_cols = detect_identity_columns(matrix[header_idx])
    fund_col = identity_cols["fund_col"]
    manager_col = identity_cols["manager_col"]

    if fund_col is None:
        raise RuntimeError("Colonne FUND non détectée.")
    if manager_col is None:
        raise RuntimeError("Colonne FUND MANAGER non détectée.")

    blocks = detect_standard_blocks(sheet_name, matrix, header_idx, date_candidates, date_bands)

    if not blocks or "volatility" in sheet_name.lower():
        matrix_blocks = detect_volatility_or_date_price_blocks(
            sheet_name, matrix, header_idx, date_candidates, date_bands
        )

        existing = {(b.valuation_date, b.start_col, b.end_col, b.unit_price_col, b.nav_col) for b in blocks}
        for b in matrix_blocks:
            key = (b.valuation_date, b.start_col, b.end_col, b.unit_price_col, b.nav_col)
            if key not in existing:
                b.block_id = len(blocks) + 1
                blocks.append(b)

    if not blocks:
        raise RuntimeError("Aucun bloc date/valorisation détecté.")

    records: List[NavRecord] = []
    current_category_raw = ""
    file_size = path.stat().st_size if path.exists() else 0
    header_text_full = " | ".join(clean_text(x) for x in matrix[header_idx] if clean_text(x))

    for row_idx, row in enumerate(matrix[header_idx + 1 :], start=header_idx + 2):
        if not row_non_empty_values(row):
            continue

        if is_category_row(row):
            current_category_raw = category_from_row(row)
            continue

        fund_name_raw = clean_text(get_cell(row, fund_col))
        fund_manager_raw = clean_text(get_cell(row, manager_col))

        if not fund_name_raw:
            continue
        if is_stop_label(fund_name_raw):
            continue
        if normalize_name_for_key(fund_name_raw) in {"FUND", "FUNDS", "FUND NAME", "NAME OF FUND"}:
            continue

        fund_name_clean = clean_display_name(fund_name_raw)
        fund_manager_clean = clean_display_name(fund_manager_raw)
        fund_name_key = normalize_name_for_key(fund_name_clean)
        fund_manager_key = normalize_name_for_key(fund_manager_clean)

        fund_category_key = normalize_name_for_key(current_category_raw)
        fund_category_fr, fund_category_conf = classify_category_fr(current_category_raw, fund_name_clean)

        for block in blocks:
            nav_value = get_cell(row, block.nav_col)
            pct_to_total = get_cell(row, block.pct_to_total_col)
            bid_price = get_cell(row, block.bid_price_col)
            offer_price = get_cell(row, block.offer_price_col)
            unit_price = get_cell(row, block.unit_price_col)
            unitholders = get_cell(row, block.unitholders_col)
            y_wtd = get_cell(row, block.yield_wtd_col)
            y_ytd = get_cell(row, block.yield_ytd_col)

            if not block.valuation_date:
                continue

            year, month, iso_week = date_parts(block.valuation_date)

            # 1. Devise de LIBELLE du fonds, deduite de son nom et de sa
            #    categorie. C est ce que l inference sait faire de fiable :
            #    « Afrinvest Dollar Fund » est bien un fonds en dollars.
            block_context = header_context_text(matrix, header_idx, block.start_col, block.end_col)
            currency_code, currency_name_fr, currency_source, currency_conf = infer_currency(
                category_raw=current_category_raw,
                fund_name=fund_name_clean,
                fund_manager=fund_manager_clean,
                header_text=header_text_full,
                block_text=block_context,
            )

            if block.block_currency_code and block.block_currency_confidence > currency_conf:
                currency_code = block.block_currency_code
                currency_name_fr = block.block_currency_name_fr
                currency_source = block.block_currency_source
                currency_conf = block.block_currency_confidence

            # 2. Choisir la colonne de prix DONT LA DEVISE est celle du fonds.
            #    Le bloc porte typiquement six colonnes de prix — Bid, Offer et
            #    Unit, en dollar et en naira. Retenir la premiere venue revenait
            #    a tirer la devise au sort : pour Afrinvest, l extracteur prenait
            #    « Offer Price (N) » = 165 509,54 et l etiquetait USD, alors que
            #    « Offer Price ($) » = 119,92 etait disponible juste a cote.
            if block.price_columns:
                _valeurs = {col: get_cell(row, col) for _, col, _ in block.price_columns}
                vl_price, vl_source, _dev_col, _prov = choose_price_column(
                    _valeurs, block.price_columns, currency_code
                )
                # Le meme bloc porte typiquement les deux devises. On les emet
                # toutes les deux, en plus de la colonne retenue : l aval choisit
                # alors sa devise sans jamais convertir. Vide quand la SEC ne
                # publie pas cette devise ce jour-la.
                vl_price_ngn, vl_price_ngn_source = price_in_currency(
                    _valeurs, block.price_columns, "NGN"
                )
                vl_price_usd, vl_price_usd_source = price_in_currency(
                    _valeurs, block.price_columns, "USD"
                )
                if _dev_col:
                    vl_currency_code = _dev_col
                    vl_currency_source = _prov
                    vl_currency_confidence = 100 if _prov.startswith("column_header") else 50
                else:
                    # Aucune colonne ne declare sa devise : on retombe sur
                    # l inference, mais en le disant.
                    vl_currency_code = currency_code
                    vl_currency_source = "inferred_" + currency_source
                    vl_currency_confidence = min(currency_conf, 50)
            else:
                # Bloc sans colonnes de prix qualifiees (chemin de repli des
                # matrices date/prix) : comportement historique conserve.
                vl_price, vl_source = choose_vl_price(offer_price, unit_price, bid_price)
                vl_currency_code = currency_code
                vl_currency_source = "inferred_" + currency_source
                vl_currency_confidence = min(currency_conf, 50)
                # Chemin de repli : aucune colonne ne declare sa devise, donc
                # aucun prix ne peut etre attribue a une devise SUR PREUVE.
                # Les laisser vides est le seul choix honnete — les remplir avec
                # `vl_price` reviendrait a affirmer une devise deduite du nom du
                # fonds, c est-a-dire le defaut meme que le lot AI a corrige.
                vl_price_ngn, vl_price_ngn_source = None, ""
                vl_price_usd, vl_price_usd_source = None, ""

            if safe_float(vl_price) is None:
                continue

            nav_currency_code = "NGN" if safe_float(nav_value) is not None else ""

            rec_dict = {
                "valuation_date": block.valuation_date,
                "fund_manager_key": fund_manager_key,
                "fund_name_key": fund_name_key,
                "fund_category_fr": fund_category_fr,
                "currency_code": currency_code,
            }
            obs_key = make_observation_key(rec_dict)

            record = NavRecord(
                source_file=path.name,
                source_url=source_url,
                source_title=source_title,
                source_page_url=source_page_url,
                source_year_page=source_year_page,
                downloaded_at_utc=downloaded_at_utc,
                file_size_bytes=file_size,

                sheet_name=sheet_name,
                source_row_number=row_idx,
                header_row_number=header_idx + 1,
                block_id=block.block_id,
                block_type=block.block_type,

                valuation_date=block.valuation_date,
                valuation_date_text=block.valuation_date_text,
                valuation_date_source=block.valuation_date_source,
                valuation_date_confidence=block.valuation_date_confidence,
                valuation_date_row=block.valuation_date_row,
                valuation_date_col=block.valuation_date_col,
                date_assignment_method=block.date_assignment_method,
                previous_or_current_hint=block.block_type,
                year=year,
                month=month,
                iso_week=iso_week,

                fund_name_raw=fund_name_raw,
                fund_name_clean=fund_name_clean,
                fund_name_key=fund_name_key,

                fund_manager_raw=fund_manager_raw,
                fund_manager_clean=fund_manager_clean,
                fund_manager_key=fund_manager_key,

                fund_category_raw=current_category_raw,
                fund_category_key=fund_category_key,
                fund_category_fr=fund_category_fr,
                fund_category_confidence=fund_category_conf,

                currency_code=currency_code,
                currency_name_fr=currency_name_fr,
                currency_source=currency_source,
                currency_confidence=currency_conf,

                nav_ngn=nav_value,
                nav_value=nav_value,
                nav_currency_code=nav_currency_code,
                pct_to_total=pct_to_total,
                bid_price=bid_price,
                offer_price=offer_price,
                unit_price=unit_price,
                vl_price=vl_price,
                vl_price_source=vl_source,
                vl_price_ngn=vl_price_ngn if vl_price_ngn is not None else "",
                vl_price_ngn_source=vl_price_ngn_source,
                vl_price_usd=vl_price_usd if vl_price_usd is not None else "",
                vl_price_usd_source=vl_price_usd_source,
                vl_currency_code=vl_currency_code,
                vl_currency_source=vl_currency_source,
                vl_currency_confidence=vl_currency_confidence,
                unitholders=unitholders,
                yield_wtd=y_wtd,
                yield_ytd=y_ytd,

                observation_key=obs_key,
                quality_flags="",
                extraction_status="OK",
            )

            record.quality_flags = validate_record_quality(record)
            records.append(record)

    return records, date_candidates, date_bands, blocks


def parse_workbook(
    path: Path,
    source_url: str = "",
    source_title: str = "",
    source_page_url: str = "",
    source_year_page: str = "",
    downloaded_at_utc: str = "",
    include_all_candidate_sheets: bool = False,
) -> Tuple[List[NavRecord], List[AuditRecord]]:
    all_records: List[NavRecord] = []
    audits: List[AuditRecord] = []

    try:
        workbook_path, workbook_prepare_status = prepare_workbook_for_openpyxl(path)
        wb = load_workbook(workbook_path, data_only=True, read_only=False)
        sheets = choose_candidate_sheets(wb, include_all_candidate_sheets=include_all_candidate_sheets)

        for sheet_name, matrix, header_idx in sheets:
            try:
                records, date_candidates, date_bands, blocks = parse_sheet_records(
                    path=path,
                    source_url=source_url,
                    source_title=source_title,
                    source_page_url=source_page_url,
                    source_year_page=source_year_page,
                    downloaded_at_utc=downloaded_at_utc,
                    sheet_name=sheet_name,
                    matrix=matrix,
                    header_idx=header_idx,
                )

                all_records.extend(records)
                dates = sorted({r.valuation_date for r in records if r.valuation_date})
                candidate_dates = sorted({c.parsed_date for c in date_candidates})
                band_desc = []
                for b in date_bands:
                    band_desc.append(f"{b.parsed_date}[{b.start_col+1}:{b.end_col+1}]")

                audits.append(
                    AuditRecord(
                        source_file=path.name,
                        source_url=source_url,
                        source_title=source_title,
                        source_page_url=source_page_url,
                        source_year_page=source_year_page,
                        status="OK",
                        message=f"Extraction réussie | workbook_prepare_status={workbook_prepare_status}",
                        sheet_name=sheet_name,
                        header_row=str(header_idx + 1),
                        detected_date_count=len(candidate_dates),
                        detected_dates=" ; ".join(candidate_dates),
                        detected_date_bands=" ; ".join(band_desc),
                        detected_block_count=len(blocks),
                        extracted_rows=len(records),
                        extracted_dates=" ; ".join(dates),
                        date_min=dates[0] if dates else "",
                        date_max=dates[-1] if dates else "",
                    )
                )

            except Exception as exc:
                audits.append(
                    AuditRecord(
                        source_file=path.name,
                        source_url=source_url,
                        source_title=source_title,
                        source_page_url=source_page_url,
                        source_year_page=source_year_page,
                        status="ERROR",
                        message=str(exc),
                        sheet_name=sheet_name,
                        header_row=str(header_idx + 1),
                    )
                )

    except Exception as exc:
        audits.append(
            AuditRecord(
                source_file=path.name,
                source_url=source_url,
                source_title=source_title,
                source_page_url=source_page_url,
                source_year_page=source_year_page,
                status="ERROR",
                message=str(exc),
            )
        )

    return all_records, audits


# =============================================================================
# 9. COHERENCE, COUVERTURE ANNUELLE, FUZZY
# =============================================================================

def make_coherence_key(record: NavRecord) -> str:
    return record.observation_key


def format_unique_values(values: Iterable[Any]) -> str:
    out = []
    seen = set()

    for v in values:
        txt = clean_text(v)
        if txt and txt not in seen:
            out.append(txt)
            seen.add(txt)

    return " ; ".join(out)


def compare_numeric_values(values: List[Any], tolerance: float) -> Tuple[bool, str]:
    nums = [safe_float(v) for v in values]
    nums = [x for x in nums if x is not None]

    if len(nums) <= 1:
        return True, ""

    if max(nums) - min(nums) <= tolerance:
        return True, ""

    return False, f"min={min(nums)} max={max(nums)} diff={max(nums)-min(nums)} tolerance={tolerance}"


def build_coherence_report(
    records: List[NavRecord],
    nav_tolerance: float = DEFAULT_NAV_TOLERANCE,
    price_tolerance: float = DEFAULT_PRICE_TOLERANCE,
) -> List[CoherenceRecord]:
    grouped: Dict[str, List[NavRecord]] = defaultdict(list)

    for r in records:
        grouped[make_coherence_key(r)].append(r)

    report: List[CoherenceRecord] = []

    for key, rows in sorted(grouped.items(), key=lambda kv: kv[0]):
        if len(rows) <= 1:
            continue

        nav_ok, nav_msg = compare_numeric_values([r.nav_value for r in rows], nav_tolerance)
        price_ok, price_msg = compare_numeric_values([r.vl_price for r in rows], price_tolerance)

        statuses = []
        statuses.append("NAV_OK" if nav_ok else "NAV_CONFLICT")
        statuses.append("VL_PRICE_OK" if price_ok else "VL_PRICE_CONFLICT")

        message_parts = []
        if nav_msg:
            message_parts.append("NAV " + nav_msg)
        if price_msg:
            message_parts.append("VL_PRICE " + price_msg)

        first = rows[0]

        report.append(
            CoherenceRecord(
                coherence_key=key,
                valuation_date=first.valuation_date,
                fund_name_key=first.fund_name_key,
                fund_manager_key=first.fund_manager_key,
                fund_category_fr=first.fund_category_fr,
                currency_code=first.currency_code,
                occurrence_count=len(rows),
                source_files=format_unique_values(r.source_file for r in rows),
                source_urls=format_unique_values(r.source_url for r in rows),
                status="|".join(statuses),
                nav_values=format_unique_values(r.nav_value for r in rows),
                vl_price_values=format_unique_values(r.vl_price for r in rows),
                message=" | ".join(message_parts),
            )
        )

    return report


def build_annual_coverage_report(
    records: List[NavRecord],
    expected_min_dates: int = 48,
    expected_max_dates: int = 54,
) -> List[CoverageRecord]:
    by_year: Dict[str, List[NavRecord]] = defaultdict(list)
    for r in records:
        if r.year:
            by_year[r.year].append(r)

    out: List[CoverageRecord] = []

    for year, rows in sorted(by_year.items()):
        dates = sorted({r.valuation_date for r in rows if r.valuation_date})
        source_files = sorted({r.source_file for r in rows if r.source_file})
        count = len(dates)

        if expected_min_dates <= count <= expected_max_dates:
            status = "OK"
            message = "Couverture annuelle cohérente avec une fréquence hebdomadaire."
        elif count < expected_min_dates:
            status = "INCOMPLETE"
            message = (
                f"Nombre de dates inférieur au seuil attendu. Attendu environ "
                f"{expected_min_dates}-{expected_max_dates}, obtenu {count}."
            )
        else:
            status = "TOO_MANY_DATES"
            message = (
                f"Nombre de dates supérieur au seuil attendu. Attendu environ "
                f"{expected_min_dates}-{expected_max_dates}, obtenu {count}. "
                f"Verifier doublons, feuilles secondaires ou dates historiques."
            )

        out.append(
            CoverageRecord(
                year=year,
                distinct_valuation_dates=count,
                first_date=dates[0] if dates else "",
                last_date=dates[-1] if dates else "",
                expected_min_dates=expected_min_dates,
                expected_max_dates=expected_max_dates,
                status=status,
                message=message,
                dates=" ; ".join(dates),
                source_files_count=len(source_files),
                records_count=len(rows),
            )
        )

    return out


def build_fuzzy_name_report(records: List[NavRecord], threshold: float = 0.94) -> List[Dict[str, Any]]:
    by_manager: Dict[str, Dict[str, str]] = defaultdict(dict)

    for r in records:
        if r.fund_manager_key and r.fund_name_key:
            by_manager[r.fund_manager_key][r.fund_name_key] = r.fund_name_clean

    report: List[Dict[str, Any]] = []

    for manager_key, funds in by_manager.items():
        keys = sorted(funds.keys())

        for i in range(len(keys)):
            for j in range(i + 1, len(keys)):
                a, b = keys[i], keys[j]

                if abs(len(a) - len(b)) > 8:
                    continue

                ratio = SequenceMatcher(None, a, b).ratio()

                if ratio >= threshold and a != b:
                    report.append(
                        {
                            "fund_manager_key": manager_key,
                            "fund_name_key_a": a,
                            "fund_name_raw_a": funds[a],
                            "fund_name_key_b": b,
                            "fund_name_raw_b": funds[b],
                            "similarity": round(ratio, 4),
                            "suggestion": "REVOIR_REFERENTIEL_MAITRE",
                        }
                    )

    return report


# =============================================================================
# 10. CSV
# =============================================================================

def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: Optional[List[str]] = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    if fieldnames is None:
        fieldnames = []
        seen = set()

        for row in rows:
            for key in row.keys():
                if key not in seen:
                    fieldnames.append(key)
                    seen.add(key)

    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


# =============================================================================
# 11. ORCHESTRATION
# =============================================================================

def run(args: argparse.Namespace) -> int:
    cache_dir = Path(args.cache_dir)

    all_records: List[NavRecord] = []
    all_audits: List[AuditRecord] = []

    # Fichiers locaux.
    for local_file in args.local_files or []:
        path = Path(local_file)

        records, audits = parse_workbook(
            path=path,
            source_url="",
            source_title=path.name,
            source_page_url="",
            source_year_page="",
            downloaded_at_utc="",
            include_all_candidate_sheets=args.include_all_candidate_sheets,
        )

        all_records.extend(records)
        all_audits.extend(audits)

    # Pages annuelles.
    for year in args.years or []:
        year = int(year)
        year_page = YEAR_URL_TEMPLATE.format(year=year)

        try:
            links = fetch_year_links(year, timeout=args.timeout)
        except Exception as exc:
            all_audits.append(
                AuditRecord(
                    source_file="",
                    source_url=year_page,
                    source_title=str(year),
                    source_page_url=year_page,
                    source_year_page=year_page,
                    status="ERROR",
                    message=f"Impossible de lire la page annuelle : {exc}",
                )
            )
            continue

        if args.max_files_per_year and args.max_files_per_year > 0:
            links = links[: args.max_files_per_year]

        for item in links:
            try:
                downloaded_at = now_iso()

                file_path = download_file(
                    item["url"],
                    cache_dir=cache_dir / str(year),
                    timeout=args.timeout,
                    sleep_seconds=args.sleep_seconds,
                )

                records, audits = parse_workbook(
                    path=file_path,
                    source_url=item["url"],
                    source_title=item["title"],
                    source_page_url=item.get("page_url", year_page),
                    source_year_page=item.get("year_page", year_page),
                    downloaded_at_utc=downloaded_at,
                    include_all_candidate_sheets=args.include_all_candidate_sheets,
                )

                all_records.extend(records)
                all_audits.extend(audits)

                ok_rows = sum(a.extracted_rows for a in audits if a.status == "OK")
                ok_dates = sorted({
                    d
                    for a in audits
                    for d in a.extracted_dates.split(" ; ")
                    if d
                })
                print(f"[OK] {year} | {file_path.name} | rows={ok_rows} | dates={','.join(ok_dates)}")

            except Exception as exc:
                all_audits.append(
                    AuditRecord(
                        source_file=safe_filename_from_url(item["url"]),
                        source_url=item["url"],
                        source_title=item["title"],
                        source_page_url=item.get("page_url", year_page),
                        source_year_page=item.get("year_page", year_page),
                        status="ERROR",
                        message=str(exc),
                    )
                )
                print(f"[ERROR] {year} | {item['url']} | {exc}", file=sys.stderr)

    records_to_write = all_records

    if args.strict_quality:
        records_to_write = [
            r for r in all_records
            if r.valuation_date
            and safe_float(r.vl_price) is not None
            and r.fund_name_key
        ]

    record_rows = [asdict(r) for r in records_to_write]
    audit_rows = [asdict(a) for a in all_audits]

    coherence_rows = [
        asdict(c)
        for c in build_coherence_report(
            records_to_write,
            nav_tolerance=args.nav_tolerance,
            price_tolerance=args.price_tolerance,
        )
    ]

    coverage_rows = [
        asdict(c)
        for c in build_annual_coverage_report(
            records_to_write,
            expected_min_dates=args.expected_min_dates,
            expected_max_dates=args.expected_max_dates,
        )
    ]

    fuzzy_rows = build_fuzzy_name_report(records_to_write, threshold=args.fuzzy_threshold)

    if args.out:
        write_csv(Path(args.out), record_rows)
    if args.audit:
        write_csv(Path(args.audit), audit_rows)
    if args.coherence:
        write_csv(Path(args.coherence), coherence_rows)
    if args.coverage:
        write_csv(Path(args.coverage), coverage_rows)
    if args.fuzzy_report:
        write_csv(Path(args.fuzzy_report), fuzzy_rows)

    print("")
    print("Extraction terminée.")
    print(f"Lignes extraites avant filtre qualité : {len(all_records)}")
    print(f"Lignes écrites : {len(record_rows)}")
    print(f"Fichiers / feuilles audités : {len(audit_rows)}")
    print(f"Lignes de cohérence inter-fichiers : {len(coherence_rows)}")
    print(f"Lignes de couverture annuelle : {len(coverage_rows)}")
    print(f"Suggestions fuzzy naming : {len(fuzzy_rows)}")
    if args.out:
        print(f"CSV données : {args.out}")
    if args.audit:
        print(f"CSV audit : {args.audit}")
    if args.coherence:
        print(f"CSV cohérence : {args.coherence}")
    if args.coverage:
        print(f"CSV couverture annuelle : {args.coverage}")
    if args.fuzzy_report:
        print(f"CSV fuzzy names : {args.fuzzy_report}")

    return 0


# =============================================================================
# 12. CLI
# =============================================================================

def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extraction fiable SEC Nigeria Weekly NAV for CIS — V5."
    )

    parser.add_argument(
        "--years",
        nargs="*",
        default=[],
        help="Années à télécharger depuis la SEC Nigeria. Exemple : --years 2026 2025 2024 2023",
    )
    parser.add_argument(
        "--local-files",
        nargs="*",
        default=[],
        help="Fichiers Excel locaux à parser.",
    )
    parser.add_argument(
        "--cache-dir",
        default="sec_ng_downloads",
        help="Dossier de cache des fichiers téléchargés.",
    )
    parser.add_argument(
        "--out",
        default="sec_ng_nav_output_v6.csv",
        help="CSV principal de sortie.",
    )
    parser.add_argument(
        "--audit",
        default="sec_ng_nav_audit_v6.csv",
        help="CSV d'audit.",
    )
    parser.add_argument(
        "--coherence",
        default="sec_ng_nav_coherence_v6.csv",
        help="CSV de cohérence inter-fichiers.",
    )
    parser.add_argument(
        "--coverage",
        default="sec_ng_nav_annual_coverage_v6.csv",
        help="CSV de couverture annuelle.",
    )
    parser.add_argument(
        "--fuzzy-report",
        default="sec_ng_nav_fuzzy_names_v6.csv",
        help="CSV de suggestions de rapprochement des noms proches.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=60,
        help="Timeout HTTP en secondes.",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=0.2,
        help="Pause entre téléchargements.",
    )
    parser.add_argument(
        "--max-files-per-year",
        type=int,
        default=0,
        help="Limiter le nombre de fichiers par année pour test. 0 = pas de limite.",
    )
    parser.add_argument(
        "--include-all-candidate-sheets",
        action="store_true",
        help="Parser toutes les feuilles candidates, y compris Volatility Measure si exploitable.",
    )
    parser.add_argument(
        "--strict-quality",
        action="store_true",
        help="Garder seulement les lignes avec date, prix de VL et nom de fonds exploitables.",
    )
    parser.add_argument(
        "--price-tolerance",
        type=float,
        default=DEFAULT_PRICE_TOLERANCE,
        help="Tolérance de comparaison du prix de VL entre fichiers.",
    )
    parser.add_argument(
        "--nav-tolerance",
        type=float,
        default=DEFAULT_NAV_TOLERANCE,
        help="Tolérance de comparaison du NAV entre fichiers.",
    )
    parser.add_argument(
        "--fuzzy-threshold",
        type=float,
        default=0.94,
        help="Seuil de similarité pour le rapport des noms proches.",
    )
    parser.add_argument(
        "--expected-min-dates",
        type=int,
        default=48,
        help="Nombre minimal attendu de dates pour une année complète.",
    )
    parser.add_argument(
        "--expected-max-dates",
        type=int,
        default=54,
        help="Nombre maximal attendu de dates pour une année complète.",
    )

    return parser


if __name__ == "__main__":
    raise SystemExit(run(build_arg_parser().parse_args()))
