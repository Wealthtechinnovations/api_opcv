#!/usr/bin/env python3
"""
CMF Tunisie OPCVM VL — Daily Incremental Scraper
=================================================

Adapted from the reference agent to the real Africafunds production schema:
- MySQL database (fund_opcvm)
- Tables: valorisations, fond_investissements, devisedechanges
- Compatible with existing import_vl_tunisie_cmf.js pipeline

Usage:
  python3 cmf_tunisie_daily.py --dry-run                    # simulation
  python3 cmf_tunisie_daily.py --production                 # import réel
  python3 cmf_tunisie_daily.py --production --lookback-days 60
  python3 cmf_tunisie_daily.py --dry-run --start-date 2026-05-19 --end-date 2026-06-01

Environment variables (from .env):
  DB_HOST, DB_USER, DB_PASSWORD, DB_NAME
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import json
import logging
import os
import re
import sys
import time
import unicodedata
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import pymysql
    pymysql.install_as_MySQLdb()
except ImportError:
    pymysql = None

try:
    from rapidfuzz import fuzz, process as rfprocess
except ImportError:
    fuzz = None
    rfprocess = None

SCRIPT_DIR = Path(__file__).resolve().parent
API_DIR = SCRIPT_DIR.parent.parent
ENV_FILE = API_DIR / ".env"

def load_dotenv():
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

load_dotenv()

DB_HOST = os.getenv("DB_HOST", "127.0.0.1")
DB_USER = os.getenv("DB_USER", "fund_opcvm")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")
DB_NAME = os.getenv("DB_NAME", "fund_opcvm")

CMF_BASE_URLS = [
    "https://www.cmf.tn/valeurs-liquidatives-des-titres-opcvm",
    "https://www.cmf.tn/?q=valeurs-liquidatives-des-titres-opcvm",
]
CMF_PAGINATION_PAGES = 9

DATA_DIR = Path(os.getenv("CMF_TUNISIE_DATA_DIR", str(API_DIR / "data" / "tunisie_cmf")))
DOWNLOAD_DIR = DATA_DIR / "downloads"
STAGING_DIR = DATA_DIR / "staging"
LOG_DIR = DATA_DIR / "logs"

REQUEST_TIMEOUT = 30
EXTREME_VARIATION_THRESHOLD = 0.20
FUZZY_THRESHOLD = 85
USER_AGENT = "AfricaFunds-CMF-Tunisie-OPCVM-Daily/2.0"

PAYS = "TUNISIE"
DEVISE = "TND"
REGULATEUR = "Conseil du Marché Financier (CMF)"
REGION = "Afrique du Nord"

CLASSIFICATION_MAP = {
    "ACTIONS": {
        "classification": "ACTIONS",
        "categorie_globale": "ACTIONS",
        "categorie_national": "ACTIONS Tunisie",
        "categorie_regional": "ACTIONS Afrique du Nord",
    },
    "OBLIGATIONS": {
        "classification": "OBLIGATIONS",
        "categorie_globale": "OBLIGATIONS",
        "categorie_national": "OBLIGATIONS Tunisie",
        "categorie_regional": "OBLIGATIONS Afrique du Nord",
    },
    "DIVERSIFIE": {
        "classification": "DIVERSIFIE",
        "categorie_globale": "DIVERSIFIE",
        "categorie_national": "DIVERSIFIE Tunisie",
        "categorie_regional": "DIVERSIFIE Afrique du Nord",
    },
    "MONETAIRE": {
        "classification": "OBLIGATIONS",
        "categorie_globale": "OBLIGATIONS",
        "categorie_national": "OBLIGATIONS Tunisie",
        "categorie_regional": "OBLIGATIONS Afrique du Nord",
    },
}

DEFAULT_CLASSIFICATION = CLASSIFICATION_MAP["OBLIGATIONS"]


def setup_logger() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("cmf_tunisie")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)
    fh = logging.FileHandler(
        LOG_DIR / f"cmf_tunisie_{datetime.now():%Y%m%d_%H%M%S}.log",
        encoding="utf-8",
    )
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger


def safe_text(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    if s.lower() in {"nan", "none", "nat", "<na>", "null", ""}:
        return ""
    return s


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def normalize_name(s: Any) -> str:
    s = strip_accents(safe_text(s)).upper()
    s = re.sub(r"[''`']+", "", s)
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def slug_key(s: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "_", normalize_name(s)).strip("_")


def to_float(x: Any) -> Optional[float]:
    s = safe_text(x)
    if not s or s.upper() in {"-", "—", "ND", "N/D", "NA"}:
        return None
    try:
        return float(s.replace("\xa0", "").replace(" ", "").replace(",", "."))
    except (ValueError, TypeError):
        return None


def to_iso_date(x: Any) -> str:
    if isinstance(x, datetime):
        return x.date().isoformat()
    if isinstance(x, date):
        return x.isoformat()
    s = safe_text(x)
    if not s:
        return ""
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except (ValueError, TypeError):
            pass
    return ""


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def detect_structure(name: str) -> Optional[str]:
    upper = (name or "").upper()
    if "SICAV" in upper:
        return "SICAV"
    if "FCP" in upper:
        return "FCP"
    return None


def get_classification(category_text: str) -> dict:
    if not category_text:
        return DEFAULT_CLASSIFICATION
    upper = category_text.upper()
    for key in CLASSIFICATION_MAP:
        if key in upper:
            return CLASSIFICATION_MAP[key]
    return DEFAULT_CLASSIFICATION


# ============================================================
# CMF WEBSITE SCRAPING
# ============================================================

def discover_cmf_files(logger: logging.Logger) -> List[dict]:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    found: Dict[str, dict] = {}

    for base_url in CMF_BASE_URLS:
        for page in range(CMF_PAGINATION_PAGES):
            page_url = base_url if page == 0 else f"{base_url}?page={page}"
            try:
                logger.info("Scraping CMF page: %s", page_url)
                resp = session.get(page_url, timeout=REQUEST_TIMEOUT)
                resp.raise_for_status()
                html = resp.text
            except Exception as exc:
                logger.warning("Error fetching %s: %s", page_url, exc)
                continue

            soup = BeautifulSoup(html, "html.parser")
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if not re.search(r"valeurs_liquidatives.*\.(xlsx?)", href, re.I):
                    continue
                url = urljoin(page_url, href)
                if url in found:
                    continue

                filename = Path(url.split("?")[0]).name
                file_date = extract_date_from_filename(filename)
                label = safe_text(a.get_text(" "))

                found[url] = {
                    "url": url,
                    "filename": filename,
                    "file_date": file_date,
                    "label": label,
                }

            time.sleep(0.3)

        if found:
            break

    logger.info("Discovered %d CMF files", len(found))
    return list(found.values())


def extract_date_from_filename(filename: str) -> str:
    m = re.search(r"(\d{6})", filename)
    if m:
        code = m.group(1)
        yy, mm, dd = int(code[:2]), int(code[2:4]), int(code[4:6])
        yyyy = 2000 + yy if yy < 80 else 1900 + yy
        try:
            return date(yyyy, mm, dd).isoformat()
        except ValueError:
            pass
    return ""


# ============================================================
# DATABASE
# ============================================================

class DbAdapter:
    def __init__(self, logger: logging.Logger):
        self.logger = logger
        self.conn = None

    def connect(self, required: bool = False):
        if not DB_PASSWORD:
            self.logger.warning("DB_PASSWORD not set — DB disabled, dry-run only")
            return
        if pymysql is None:
            if required:
                raise RuntimeError("pymysql not installed")
            self.logger.warning("pymysql not installed — DB disabled")
            return
        try:
            self.conn = pymysql.connect(
                host=DB_HOST,
                user=DB_USER,
                password=DB_PASSWORD,
                database=DB_NAME,
                charset="utf8mb4",
                autocommit=False,
            )
            self.logger.info("Connected to MySQL %s@%s/%s", DB_USER, DB_HOST, DB_NAME)
        except Exception as exc:
            if required:
                raise
            self.logger.warning("DB connection failed (non-fatal in dry-run): %s", exc)

    def close(self):
        if self.conn:
            self.conn.close()

    def ensure_audit_tables(self):
        if not self.conn:
            return
        with self.conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cmf_import_audit (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    import_batch VARCHAR(100) NOT NULL,
                    run_mode VARCHAR(20) NOT NULL,
                    files_discovered INT DEFAULT 0,
                    files_downloaded INT DEFAULT 0,
                    nav_parsed INT DEFAULT 0,
                    nav_imported INT DEFAULT 0,
                    nav_skipped INT DEFAULT 0,
                    nav_extreme INT DEFAULT 0,
                    new_funds INT DEFAULT 0,
                    dividends_found INT DEFAULT 0,
                    errors TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cmf_extreme_variations (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    fund_id INT,
                    fund_name VARCHAR(255),
                    vl_date DATE NOT NULL,
                    vl_new DOUBLE,
                    vl_previous DOUBLE,
                    variation_pct DOUBLE,
                    source_file VARCHAR(255),
                    import_batch VARCHAR(100),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_fund_date (fund_id, vl_date)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS cmf_new_funds_queue (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    fund_name_source VARCHAR(255) NOT NULL,
                    fund_key VARCHAR(255) NOT NULL,
                    manager VARCHAR(255),
                    category VARCHAR(100),
                    first_date DATE,
                    vl_value DOUBLE,
                    status VARCHAR(30) DEFAULT 'PENDING',
                    import_batch VARCHAR(100),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE KEY uk_fund_key (fund_key)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        self.conn.commit()
        self.logger.info("Audit tables verified/created")

    def fetch_existing_tunisie_dates(self) -> set:
        if not self.conn:
            return set()
        with self.conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT DATE_FORMAT(v.date, '%%Y-%%m-%%d')
                FROM valorisations v
                JOIN fond_investissements f ON v.fund_id = f.id
                WHERE f.pays = %s
            """, (PAYS,))
            return {r[0] for r in cur.fetchall() if r[0]}

    def fetch_tunisie_funds(self) -> List[dict]:
        if not self.conn:
            return []
        with self.conn.cursor(pymysql.cursors.DictCursor) as cur:
            cur.execute("""
                SELECT id, nom_fond, code_ISIN, societe_gestion,
                       classification, categorie_globale, categorie_national,
                       categorie_regional, periodicite, affectation,
                       structure_fond, indice_benchmark,
                       date_premiere_vl, datejour
                FROM fond_investissements
                WHERE pays = %s
            """, (PAYS,))
            return cur.fetchall()

    def fetch_exchange_rates(self) -> Tuple[dict, dict]:
        if not self.conn:
            return {}, {}
        eur_map, usd_map = {}, {}
        with self.conn.cursor() as cur:
            cur.execute("""
                SELECT paire, DATE_FORMAT(date, '%%Y-%%m-%%d') as d, value
                FROM devisedechanges
                WHERE paire IN ('EUR/TND', 'USD/TND') AND value > 0
                ORDER BY date
            """)
            for paire, d, value in cur.fetchall():
                if paire == "EUR/TND":
                    eur_map[d] = float(value)
                else:
                    usd_map[d] = float(value)
        return eur_map, usd_map

    def fetch_existing_vl_keys(self, fund_ids: List[int]) -> set:
        if not self.conn or not fund_ids:
            return set()
        keys = set()
        batch_size = 500
        for i in range(0, len(fund_ids), batch_size):
            batch = fund_ids[i : i + batch_size]
            placeholders = ",".join(["%s"] * len(batch))
            with self.conn.cursor() as cur:
                cur.execute(
                    f"SELECT fund_id, DATE_FORMAT(date, '%%Y-%%m-%%d') FROM valorisations WHERE fund_id IN ({placeholders})",
                    batch,
                )
                for fid, d in cur.fetchall():
                    keys.add(f"{fid}|||{d}")
        return keys


def get_rate(rate_map: dict, target_date: str) -> Optional[float]:
    if not rate_map:
        return None
    if target_date in rate_map:
        return rate_map[target_date]
    sorted_dates = sorted(rate_map.keys())
    for d in reversed(sorted_dates):
        if d <= target_date:
            return rate_map[d]
    return rate_map[sorted_dates[0]] if sorted_dates else None


# ============================================================
# EXCEL PARSING
# ============================================================

def read_excel_sheets(path: Path) -> List[Tuple[str, List[List[Any]]]]:
    data = path.read_bytes()
    if data[:2] == b"PK":
        if openpyxl is None:
            raise RuntimeError("openpyxl required for .xlsx")
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
        sheets = [
            (ws.title, [list(r) for r in ws.iter_rows(values_only=True)])
            for ws in wb.worksheets
        ]
        wb.close()
        return sheets
    if data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        if xlrd is None:
            raise RuntimeError("xlrd required for .xls")
        book = xlrd.open_workbook(file_contents=data)
        sheets = []
        for sh in book.sheets():
            rows = []
            for r in range(sh.nrows):
                vals = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            vals.append(datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode)))
                        except Exception:
                            vals.append(cell.value)
                    else:
                        vals.append(cell.value)
                rows.append(vals)
            sheets.append((sh.name, rows))
        return sheets
    import pandas as pd
    if b"<table" in data[:5000].lower() or b"<html" in data[:5000].lower():
        return [
            (f"html_table_{i + 1}", df.values.tolist())
            for i, df in enumerate(pd.read_html(BytesIO(data)))
        ]
    raise ValueError(f"Unrecognized format: {path}")


def parse_cmf_excel(logger: logging.Logger, file_path: Path, file_date: str) -> dict:
    result = {
        "nav_rows": [],
        "dividends": [],
        "errors": [],
        "fund_count": 0,
    }

    if not file_date:
        result["errors"].append(f"No date for file {file_path.name}")
        return result

    try:
        sheets = read_excel_sheets(file_path)
    except Exception as exc:
        result["errors"].append(f"Read error {file_path.name}: {exc}")
        return result

    for sheet_name, rows in sheets:
        section_context = {
            "affectation": "",
            "categorie": "",
            "periodicite": "",
            "forme": "",
            "is_distribution": False,
            "has_dividends": False,
        }

        for row_num, raw_row in enumerate(rows, 1):
            row = list(raw_row) + [None] * 20

            row_text = " ".join(
                safe_text(c).upper() for c in row[:9] if safe_text(c)
            )

            if "DISTRIBUTION" in row_text:
                section_context["is_distribution"] = True
                section_context["has_dividends"] = True
            if "CAPITALISATION" in row_text:
                section_context["is_distribution"] = False
                section_context["has_dividends"] = False

            if "OBLIGAT" in row_text and not _is_fund_row(row):
                section_context["categorie"] = "OBLIGATIONS"
            elif "ACTION" in row_text and not _is_fund_row(row):
                section_context["categorie"] = "ACTIONS"
            elif "MIX" in row_text and not _is_fund_row(row):
                section_context["categorie"] = "DIVERSIFIE"
            elif "MONETAIRE" in row_text and not _is_fund_row(row):
                section_context["categorie"] = "MONETAIRE"

            if "HEBDOM" in row_text and not _is_fund_row(row):
                section_context["periodicite"] = "HEBDOMADAIRE"
            elif "QUOTID" in row_text and not _is_fund_row(row):
                section_context["periodicite"] = "QUOTIDIENNE"

            if "SICAV" in row_text and not _is_fund_row(row):
                section_context["forme"] = "SICAV"
            elif "FCP" in row_text and not _is_fund_row(row):
                section_context["forme"] = "FCP"

            if not _is_fund_row(row):
                continue

            fund_name = safe_text(row[1])
            manager = safe_text(row[2])
            open_date = to_iso_date(row[3])

            if section_context["has_dividends"]:
                div_date = to_iso_date(row[4])
                div_amount = to_float(row[5])
                vl_prev = to_float(row[7])
                vl_current = to_float(row[8])
            else:
                div_date = ""
                div_amount = None
                vl_prev = to_float(row[7])
                vl_current = to_float(row[8])

            if vl_current is None or vl_current <= 0:
                continue

            variation_vl = None
            variation_pct = None
            if vl_prev and vl_prev > 0:
                variation_vl = vl_current - vl_prev
                variation_pct = variation_vl / vl_prev

            nav_entry = {
                "date": file_date,
                "fund_name_source": fund_name,
                "fund_name_normalized": normalize_name(fund_name),
                "fund_key": slug_key(fund_name),
                "manager": manager,
                "open_date": open_date,
                "vl": vl_current,
                "vl_prev": vl_prev,
                "variation_vl": variation_vl,
                "variation_pct": variation_pct,
                "categorie": section_context["categorie"],
                "periodicite": section_context["periodicite"],
                "affectation": "DISTRIBUANT" if section_context["is_distribution"] else "CAPITALISANT",
                "forme": section_context["forme"] or detect_structure(fund_name),
                "source_file": file_path.name,
                "sheet": sheet_name,
                "row_num": row_num,
            }
            result["nav_rows"].append(nav_entry)

            if div_date or (div_amount and div_amount > 0):
                result["dividends"].append({
                    "fund_name": fund_name,
                    "fund_key": slug_key(fund_name),
                    "date": file_date,
                    "div_date": div_date,
                    "div_amount": div_amount or 0,
                })

    result["fund_count"] = len(result["nav_rows"])
    logger.info(
        "Parsed %s: %d funds, %d dividends, %d errors",
        file_path.name,
        result["fund_count"],
        len(result["dividends"]),
        len(result["errors"]),
    )
    return result


def _is_fund_row(row: list) -> bool:
    try:
        idx = int(float(row[0]))
        if idx < 1 or idx > 500:
            return False
    except (ValueError, TypeError):
        return False
    name = safe_text(row[1])
    return len(name) >= 3


# ============================================================
# FUND MATCHING
# ============================================================

def match_funds(
    logger: logging.Logger,
    nav_rows: List[dict],
    existing_funds: List[dict],
) -> Tuple[Dict[str, dict], List[dict]]:
    prod_by_norm = {}
    for f in existing_funds:
        norm = normalize_name(f["nom_fond"])
        prod_by_norm[norm] = f

    matching = {}
    new_funds = []
    seen_new_keys = set()

    for nav in nav_rows:
        key = nav["fund_key"]
        if key in matching:
            continue

        cmf_norm = nav["fund_name_normalized"]

        if cmf_norm in prod_by_norm:
            f = prod_by_norm[cmf_norm]
            matching[key] = {"fund_id": f["id"], "fund_name": f["nom_fond"], "method": "EXACT"}
            continue

        found = False
        for pnorm, pf in prod_by_norm.items():
            if cmf_norm in pnorm or pnorm in cmf_norm:
                matching[key] = {"fund_id": pf["id"], "fund_name": pf["nom_fond"], "method": "PARTIAL"}
                found = True
                break
        if found:
            continue

        if rfprocess and fuzz:
            choices = list(prod_by_norm.keys())
            if choices:
                m = rfprocess.extractOne(cmf_norm, choices, scorer=fuzz.WRatio)
                if m and m[1] >= FUZZY_THRESHOLD:
                    pf = prod_by_norm[m[0]]
                    matching[key] = {
                        "fund_id": pf["id"],
                        "fund_name": pf["nom_fond"],
                        "method": f"FUZZY({m[1]:.0f}%)",
                    }
                    continue
        else:
            best_score = 0
            best_fund = None
            for pnorm, pf in prod_by_norm.items():
                score = _dice_similarity(cmf_norm, pnorm)
                if score > best_score:
                    best_score = score
                    best_fund = pf
            if best_score >= 0.85:
                matching[key] = {
                    "fund_id": best_fund["id"],
                    "fund_name": best_fund["nom_fond"],
                    "method": f"FUZZY({best_score*100:.0f}%)",
                }
                continue

        if key not in seen_new_keys:
            new_funds.append(nav)
            seen_new_keys.add(key)

    methods = {}
    for v in matching.values():
        m = v["method"].split("(")[0]
        methods[m] = methods.get(m, 0) + 1
    logger.info("Fund matching: %s matched, %s new — %s", len(matching), len(new_funds), methods)

    return matching, new_funds


def _dice_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0
    if a == b:
        return 1
    bg_a = {a[i : i + 2] for i in range(len(a) - 1)}
    bg_b = {b[i : i + 2] for i in range(len(b) - 1)}
    intersection = len(bg_a & bg_b)
    return (2 * intersection) / (len(bg_a) + len(bg_b)) if (bg_a or bg_b) else 0


# ============================================================
# QUALITY CONTROL
# ============================================================

def quality_control(
    logger: logging.Logger,
    nav_rows: List[dict],
    matching: Dict[str, dict],
    db: DbAdapter,
    eur_rates: dict,
    usd_rates: dict,
) -> dict:
    result = {
        "clean": [],
        "extreme": [],
        "no_match": [],
        "no_rate": [],
        "duplicates": 0,
    }

    fund_ids = list({m["fund_id"] for m in matching.values()})
    existing_keys = db.fetch_existing_vl_keys(fund_ids)

    seen = set()
    for nav in nav_rows:
        key = nav["fund_key"]
        m = matching.get(key)
        if not m:
            result["no_match"].append(nav)
            continue

        fund_id = m["fund_id"]
        vl_date = nav["date"]
        dedup_key = f"{fund_id}|||{vl_date}"

        if dedup_key in seen:
            result["duplicates"] += 1
            continue
        seen.add(dedup_key)

        if dedup_key in existing_keys:
            result["duplicates"] += 1
            continue

        if nav["variation_pct"] is not None and abs(nav["variation_pct"]) > EXTREME_VARIATION_THRESHOLD:
            result["extreme"].append({**nav, "fund_id": fund_id, "fund_name_prod": m["fund_name"]})
            continue

        eur_rate = get_rate(eur_rates, vl_date)
        usd_rate = get_rate(usd_rates, vl_date)
        if not eur_rate or not usd_rate:
            result["no_rate"].append(nav)
            continue

        nav["fund_id"] = fund_id
        nav["fund_name_prod"] = m["fund_name"]
        nav["value_eur"] = nav["vl"] / eur_rate
        nav["value_usd"] = nav["vl"] / usd_rate
        nav["eur_rate"] = eur_rate
        nav["usd_rate"] = usd_rate
        result["clean"].append(nav)

    logger.info(
        "Quality: %d clean, %d extreme, %d no_match, %d no_rate, %d duplicates",
        len(result["clean"]),
        len(result["extreme"]),
        len(result["no_match"]),
        len(result["no_rate"]),
        result["duplicates"],
    )
    return result


# ============================================================
# IMPORT
# ============================================================

def import_to_db(
    logger: logging.Logger,
    db: DbAdapter,
    clean_rows: List[dict],
    extreme_rows: List[dict],
    new_funds: List[dict],
    import_batch: str,
    dry_run: bool,
) -> dict:
    stats = {
        "inserted": 0,
        "extreme_logged": 0,
        "new_funds_queued": 0,
        "metadata_updated": 0,
    }

    if dry_run or not db.conn:
        return stats

    try:
        with db.conn.cursor() as cur:
            for nav in clean_rows:
                fund_id = nav["fund_id"]
                fund_name = nav["fund_name_prod"]
                vl = nav["vl"]
                vl_date = nav["date"]
                value_eur = nav["value_eur"]
                value_usd = nav["value_usd"]
                benchmark = "Tunindex"

                cur.execute(
                    """INSERT INTO valorisations
                       (fund_id, fund_name, value, value_USD, value_EUR,
                        dividende, dividende_EUR, dividende_USD,
                        vl_ajuste, vl_ajuste_EUR, vl_ajuste_USD,
                        indice_name, base_100, base_100_InRef, tsr, tra,
                        indRef, indRef_EUR, indRef_USD,
                        indice_comparaison, actif_net, actif_net_USD, actif_net_EUR,
                        libelle_fond, souscription, ID_indice, rachat, date)
                       VALUES (%s,%s,%s,%s,%s, 0,0,0, %s,%s,%s,
                               %s,0,0,0,0, 0,0,0, 0,0,0,0,
                               %s,0,%s,0,%s)""",
                    (
                        fund_id, fund_name, vl, value_usd, value_eur,
                        vl, value_eur, value_usd,
                        benchmark,
                        fund_name, benchmark, vl_date,
                    ),
                )
                stats["inserted"] += 1

            for ext in extreme_rows:
                cur.execute(
                    """INSERT INTO cmf_extreme_variations
                       (fund_id, fund_name, vl_date, vl_new, vl_previous,
                        variation_pct, source_file, import_batch)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        ext.get("fund_id"),
                        ext.get("fund_name_source", ""),
                        ext["date"],
                        ext["vl"],
                        ext.get("vl_prev"),
                        ext.get("variation_pct"),
                        ext.get("source_file", ""),
                        import_batch,
                    ),
                )
                stats["extreme_logged"] += 1

            for nf in new_funds:
                try:
                    cur.execute(
                        """INSERT INTO cmf_new_funds_queue
                           (fund_name_source, fund_key, manager, category,
                            first_date, vl_value, import_batch)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)
                           ON DUPLICATE KEY UPDATE
                           first_date = LEAST(first_date, VALUES(first_date)),
                           vl_value = VALUES(vl_value)""",
                        (
                            nf["fund_name_source"],
                            nf["fund_key"],
                            nf.get("manager", ""),
                            nf.get("categorie", ""),
                            nf["date"],
                            nf["vl"],
                            import_batch,
                        ),
                    )
                    stats["new_funds_queued"] += 1
                except Exception:
                    pass

            affected_fund_ids = list({nav["fund_id"] for nav in clean_rows})
            for fid in affected_fund_ids:
                cur.execute(
                    """UPDATE fond_investissements f
                       SET datejour = (SELECT MAX(date) FROM valorisations WHERE fund_id = %s),
                           date_premiere_vl = COALESCE(
                               (SELECT MIN(date) FROM valorisations WHERE fund_id = %s),
                               f.date_premiere_vl
                           )
                       WHERE f.id = %s""",
                    (fid, fid, fid),
                )
                stats["metadata_updated"] += 1

        db.conn.commit()
        logger.info("Import committed: %s", stats)

    except Exception as exc:
        db.conn.rollback()
        logger.error("Import rolled back: %s", exc)
        raise

    return stats


# ============================================================
# LOCKFILE
# ============================================================

@contextlib.contextmanager
def lockfile(path: Path):
    if path.exists():
        pid_text = path.read_text().strip()
        try:
            pid = int(pid_text)
            os.kill(pid, 0)
            raise RuntimeError(f"Lock held by PID {pid}: {path}")
        except (ValueError, ProcessLookupError, PermissionError):
            path.unlink(missing_ok=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(str(os.getpid()), encoding="utf-8")
    try:
        yield
    finally:
        with contextlib.suppress(Exception):
            path.unlink()


# ============================================================
# REPORT
# ============================================================

def write_report(
    logger: logging.Logger,
    summary: dict,
    clean: List[dict],
    extreme: List[dict],
    new_funds: List[dict],
    no_match: List[dict],
) -> Path:
    STAGING_DIR.mkdir(parents=True, exist_ok=True)
    report_path = STAGING_DIR / f"cmf_tunisie_report_{datetime.now():%Y%m%d_%H%M%S}.json"
    report = {
        "summary": summary,
        "clean_count": len(clean),
        "extreme_count": len(extreme),
        "new_funds_count": len(new_funds),
        "no_match_count": len(no_match),
        "extreme_details": extreme[:50],
        "new_funds_details": [
            {"name": nf["fund_name_source"], "key": nf["fund_key"], "manager": nf.get("manager", "")}
            for nf in new_funds
        ],
        "no_match_details": [
            {"name": nm["fund_name_source"], "key": nm["fund_key"]}
            for nm in no_match[:50]
        ],
        "clean_dates": sorted(set(c["date"] for c in clean)),
        "clean_funds_sample": [
            {"fund": c["fund_name_prod"], "date": c["date"], "vl": c["vl"]}
            for c in clean[:20]
        ],
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    logger.info("Report written: %s", report_path)
    return report_path


# ============================================================
# MAIN
# ============================================================

def main(argv: Optional[Sequence[str]] = None) -> None:
    ap = argparse.ArgumentParser(description="CMF Tunisie OPCVM Daily Scraper")
    mode = ap.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true", help="Simulation only")
    mode.add_argument("--production", action="store_true", help="Import to production DB")
    ap.add_argument("--lookback-days", type=int, default=45, help="Max lookback window")
    ap.add_argument("--start-date", default="", help="Filter: only files >= this date (YYYY-MM-DD)")
    ap.add_argument("--end-date", default="", help="Filter: only files <= this date (YYYY-MM-DD)")
    ap.add_argument("--skip-scrape", action="store_true", help="Skip web scraping, use existing downloads")
    args = ap.parse_args(argv)

    logger = setup_logger()
    import_batch = f"cmf_tunisie_{datetime.now():%Y%m%d_%H%M%S}"
    logger.info("=== CMF Tunisie Daily Scraper — %s ===", "DRY-RUN" if args.dry_run else "PRODUCTION")
    logger.info("Import batch: %s", import_batch)

    with lockfile(LOG_DIR / "cmf_tunisie_daily.lock"):
        db = DbAdapter(logger)
        db.connect(required=args.production)
        try:
            if args.production:
                db.ensure_audit_tables()

            existing_dates = db.fetch_existing_tunisie_dates()
            logger.info("Existing dates in DB: %d", len(existing_dates))

            existing_funds = db.fetch_tunisie_funds()
            logger.info("Existing Tunisie funds: %d", len(existing_funds))

            eur_rates, usd_rates = db.fetch_exchange_rates()
            logger.info("Exchange rates: EUR/TND=%d days, USD/TND=%d days", len(eur_rates), len(usd_rates))

            # Step 1: Discover files
            if not args.skip_scrape:
                cmf_files = discover_cmf_files(logger)
            else:
                cmf_files = _local_files_only(logger)

            # Step 2: Filter and download
            today = date.today()
            min_date = today - timedelta(days=args.lookback_days)
            downloaded = []

            session = requests.Session()
            session.headers.update({"User-Agent": USER_AGENT})
            DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

            for cf in cmf_files:
                fd = cf["file_date"]
                if not fd:
                    logger.debug("Skipping file without date: %s", cf["filename"])
                    continue
                if args.start_date and fd < args.start_date:
                    continue
                if args.end_date and fd > args.end_date:
                    continue
                try:
                    fd_date = datetime.strptime(fd, "%Y-%m-%d").date()
                except ValueError:
                    continue
                if fd_date < min_date:
                    continue

                if fd in existing_dates:
                    logger.debug("Date %s already in DB, skipping", fd)
                    continue

                local_path = DOWNLOAD_DIR / cf["filename"]
                if not local_path.exists():
                    try:
                        logger.info("Downloading %s ...", cf["filename"])
                        resp = session.get(cf["url"], timeout=REQUEST_TIMEOUT)
                        resp.raise_for_status()
                        local_path.write_bytes(resp.content)
                        time.sleep(0.3)
                    except Exception as exc:
                        logger.warning("Download error %s: %s", cf["url"], exc)
                        continue

                cf["local_path"] = str(local_path)
                downloaded.append(cf)

            logger.info("Files to process: %d (out of %d discovered)", len(downloaded), len(cmf_files))

            # Step 3: Parse all files
            all_nav = []
            all_dividends = []
            all_errors = []

            for cf in downloaded:
                parsed = parse_cmf_excel(logger, Path(cf["local_path"]), cf["file_date"])
                all_nav.extend(parsed["nav_rows"])
                all_dividends.extend(parsed["dividends"])
                all_errors.extend(parsed["errors"])

            logger.info("Total parsed: %d NAV rows, %d dividends, %d errors", len(all_nav), len(all_dividends), len(all_errors))

            if not all_nav:
                logger.info("No new NAV data to import. Done.")
                summary = {
                    "status": "NO_NEW_DATA",
                    "discovered": len(cmf_files),
                    "downloaded": len(downloaded),
                    "parsed": 0,
                }
                print(json.dumps(summary, indent=2))
                return

            # Step 4: Match funds
            matching, new_funds = match_funds(logger, all_nav, existing_funds)

            # Step 5: Quality control
            qc = quality_control(logger, all_nav, matching, db, eur_rates, usd_rates)

            # Step 6: Import
            import_stats = import_to_db(
                logger, db,
                qc["clean"], qc["extreme"], new_funds,
                import_batch, args.dry_run,
            )

            # Step 7: Audit log
            summary = {
                "status": "DRY_RUN_OK" if args.dry_run else "PRODUCTION_OK",
                "import_batch": import_batch,
                "discovered_files": len(cmf_files),
                "downloaded_files": len(downloaded),
                "parsed_nav": len(all_nav),
                "clean_nav": len(qc["clean"]),
                "extreme_nav": len(qc["extreme"]),
                "no_match": len(qc["no_match"]),
                "no_rate": len(qc["no_rate"]),
                "duplicates": qc["duplicates"],
                "new_funds": len(new_funds),
                "dividends": len(all_dividends),
                "errors": len(all_errors),
                **import_stats,
                "dates_imported": sorted(set(c["date"] for c in qc["clean"])),
            }

            if db.conn and not args.dry_run:
                try:
                    with db.conn.cursor() as cur:
                        cur.execute(
                            """INSERT INTO cmf_import_audit
                               (import_batch, run_mode, files_discovered, files_downloaded,
                                nav_parsed, nav_imported, nav_skipped, nav_extreme,
                                new_funds, dividends_found, errors)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (
                                import_batch, "PRODUCTION",
                                len(cmf_files), len(downloaded),
                                len(all_nav), import_stats["inserted"],
                                qc["duplicates"], len(qc["extreme"]),
                                len(new_funds), len(all_dividends),
                                json.dumps(all_errors[:10], default=str) if all_errors else None,
                            ),
                        )
                    db.conn.commit()
                except Exception as exc:
                    logger.warning("Audit log insert failed: %s", exc)

            report_path = write_report(logger, summary, qc["clean"], qc["extreme"], new_funds, qc["no_match"])
            summary["report_path"] = str(report_path)

            logger.info("=== RESULT ===")
            print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))

        finally:
            db.close()


def _local_files_only(logger: logging.Logger) -> List[dict]:
    files = []
    if DOWNLOAD_DIR.exists():
        for p in sorted(DOWNLOAD_DIR.glob("valeurs_liquidatives_*.xls*")):
            fd = extract_date_from_filename(p.name)
            files.append({
                "url": "",
                "filename": p.name,
                "file_date": fd,
                "label": "",
                "local_path": str(p),
            })
    logger.info("Local files found: %d", len(files))
    return files


if __name__ == "__main__":
    main()
