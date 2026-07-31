#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SEC Nigeria — Chargeur du classeur d'extraction officielle 2011-2026
=====================================================================

Charge le classeur `Nigeria_SEC_OPCVM_NAV_2011_2026.xlsx` (extraction complete
et normalisee des 686 publications officielles SEC Nigeria) dans des tables de
STAGING additives, resout l'identite des fonds, puis produit un rapport de
comparaison avec la production — SANS ECRIRE EN PRODUCTION par defaut.

Conforme a PROMPT_CLAUDE_CODE_NIGERIA_OPCVM_ZERO_REGRESSION_V2_2.md :
  - grain = fonds x date de valorisation (date du BLOC, pas du fichier) ;
  - actif net total, VL/Unit Price, Bid et Offer restent des mesures DISTINCTES ;
  - NGN et USD restent des devises distinctes, aucune conversion implicite ;
  - aucune valeur inventee, aucune imputation, aucun decalage de date en masse ;
  - zeros publies conserves (SOURCE_ZERO), jamais convertis en NULL ;
  - provenance complete conservee (document SEC, fichier, URL, date du rapport) ;
  - dry-run par defaut ; ecriture staging seulement avec --execute ;
  - promotion en production JAMAIS faite par ce script (voir --report puis
    l'etape de promotion dediee, apres validation humaine explicite).

Tables additives creees (prefixe `sec_ng_`, zero collision avec l'existant) :
  sec_ng_observations   — 1 ligne par fonds x date, toutes mesures + provenance
  sec_ng_fund_aliases   — resolution nom SEC -> fund_id, avec confiance/statut
  sec_ng_load_logs      — journal des chargements

Usage :
  python3 sec_ng_xlsx_loader.py --xlsx <fichier.xlsx> --dry-run        # defaut
  python3 sec_ng_xlsx_loader.py --xlsx <fichier.xlsx> --execute        # staging
  python3 sec_ng_xlsx_loader.py --xlsx <fichier.xlsx> --report         # comparaison prod
  python3 sec_ng_xlsx_loader.py --selftest                             # tests internes

Dependances : openpyxl, pymysql, rapidfuzz (optionnel, ameliore le matching)
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import pymysql
except ImportError:
    pymysql = None
try:
    from rapidfuzz import fuzz
except ImportError:
    fuzz = None

SCRIPT_DIR = Path(__file__).resolve().parent
API_DIR = SCRIPT_DIR.parent.parent
REPORT_DIR = API_DIR / "data" / "sec_ng_xlsx" / "reports"

PAYS = "NIGERIA"
SHEET_DATA = "Données"

# Colonnes attendues du classeur (verifiees sur le fichier reel du 2026-07-22).
# Si une colonne manque -> SCHEMA_DRIFT, arret : on ne lit JAMAIS par position.
EXPECTED_COLUMNS = [
    "Date de valorisation", "Fonds", "Gestionnaire", "Catégorie SEC",
    "Actif net total (NGN)", "Actif net total (USD)", "VL / Unit Price (NGN)",
    "Bid Price (NGN)", "Bid Price (USD)", "Offer Price (NGN)", "Offer Price (USD)",
    "Statut qualité", "Note qualité", "Conflit source", "Date du rapport source",
    "Document SEC ID", "Fichier source", "URL source",
]

log = logging.getLogger("sec_ng_xlsx")


# ---------------------------------------------------------------------------
# Normalisation — jamais destructive, jamais inventive
# ---------------------------------------------------------------------------
def normalize_name(s):
    """NFKD, sans accents, majuscules, & -> AND, ponctuation neutralisee."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.upper().replace("&", " AND ")
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def compact_key(s):
    """Clef compacte sans espaces : rapproche 'CanaryGrowth' et 'Canary Growth'."""
    return normalize_name(s).replace(" ", "")


def to_iso(v):
    """Date Excel/py -> 'YYYY-MM-DD'. Jamais de date fabriquee."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else None


def to_num(v):
    """Valeur numerique native Excel conservee telle quelle.

    Le classeur source est DEJA normalise (openpyxl rend des float natifs) :
    on ne re-parse donc PAS la chaine affichee, conformement a la regle
    « si une valeur est deja numerique dans Excel, conserve sa valeur native ».
    Les rares chaines residuelles sont traitees avec detection explicite du
    separateur decimal ; toute ambiguite renvoie None (jamais de supposition).
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.upper() in ("N/A", "NA", "NIL", "NONE", "ND", "-", "--"):
        return None
    neg = s.startswith("(") and s.endswith(")")          # negatif comptable
    if neg:
        s = s[1:-1]
    s = re.sub(r"[₦$]|NGN|USD", "", s, flags=re.I).strip()
    s = s.replace(" ", "").replace(" ", "").replace("'", "").replace("’", "")
    has_dot, has_com = "." in s, "," in s
    if has_dot and has_com:                              # le DERNIER separateur decide
        dec = "." if s.rfind(".") > s.rfind(",") else ","
        s = s.replace(",", "") if dec == "." else s.replace(".", "").replace(",", ".")
    elif has_com:
        parts = s.split(",")
        # ',' est un separateur de milliers si tous les groupes suivants font 3 chiffres
        s = s.replace(",", "") if all(len(p) == 3 for p in parts[1:]) and len(parts) > 1 \
            else s.replace(",", ".")
    s = s.replace(" ", "")
    try:
        val = float(s)
    except ValueError:
        return None
    return -val if neg else val


# ---------------------------------------------------------------------------
# Base de donnees
# ---------------------------------------------------------------------------
def load_env():
    env_path = API_DIR / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def db_connect():
    if pymysql is None:
        raise RuntimeError("pymysql non installe")
    load_env()
    return pymysql.connect(
        host=os.environ.get("DB_HOST", "127.0.0.1"),
        user=os.environ.get("DB_USER", "fund_opcvm"),
        password=os.environ.get("DB_PASSWORD", ""),
        database=os.environ.get("DB_NAME", "fund_opcvm"),
        charset="utf8mb4", autocommit=True,
        cursorclass=pymysql.cursors.DictCursor,
    )


DDL = [
    """CREATE TABLE IF NOT EXISTS sec_ng_observations (
        id INT AUTO_INCREMENT PRIMARY KEY,
        valuation_date DATE NOT NULL,
        fund_name_raw VARCHAR(255) NOT NULL,
        fund_key VARCHAR(255) NOT NULL,
        manager_raw VARCHAR(255) NULL,
        category_sec VARCHAR(255) NULL,
        net_assets_ngn DECIMAL(28,6) NULL,
        net_assets_usd DECIMAL(28,6) NULL,
        unit_price_ngn DECIMAL(20,6) NULL,
        bid_price_ngn DECIMAL(20,6) NULL,
        bid_price_usd DECIMAL(20,6) NULL,
        offer_price_ngn DECIMAL(20,6) NULL,
        offer_price_usd DECIMAL(20,6) NULL,
        quality_status VARCHAR(20) NOT NULL,
        quality_note TEXT NULL,
        has_conflict TINYINT NOT NULL DEFAULT 0,
        report_date DATE NULL,
        sec_document_id VARCHAR(32) NULL,
        source_file VARCHAR(255) NULL,
        source_url VARCHAR(512) NULL,
        row_hash CHAR(40) NOT NULL,
        matched_fund_id INT NULL,
        match_status VARCHAR(20) NOT NULL DEFAULT 'PENDING',
        match_confidence DOUBLE NULL,
        compare_status VARCHAR(24) NULL,
        load_batch VARCHAR(40) NULL,
        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        UNIQUE KEY uq_row (row_hash),
        KEY idx_date (valuation_date),
        KEY idx_key (fund_key),
        KEY idx_match (match_status),
        KEY idx_fund (matched_fund_id),
        KEY idx_compare (compare_status)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
    """CREATE TABLE IF NOT EXISTS sec_ng_fund_aliases (
        id INT AUTO_INCREMENT PRIMARY KEY,
        fund_key VARCHAR(255) NOT NULL,
        fund_name_raw VARCHAR(255) NOT NULL,
        normalized_name VARCHAR(255) NOT NULL,
        compact_name VARCHAR(255) NOT NULL,
        manager_raw VARCHAR(255) NULL,
        fund_id INT NULL,
        match_status VARCHAR(20) NOT NULL,
        confidence DOUBLE NULL,
        review_reason VARCHAR(255) NULL,
        first_seen DATE NULL,
        last_seen DATE NULL,
        observations INT NULL,
        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        UNIQUE KEY uq_key (fund_key),
        KEY idx_status (match_status),
        KEY idx_fund (fund_id)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
    """CREATE TABLE IF NOT EXISTS sec_ng_load_logs (
        id INT AUTO_INCREMENT PRIMARY KEY,
        batch VARCHAR(40) NOT NULL,
        started_at DATETIME NOT NULL,
        finished_at DATETIME NULL,
        status VARCHAR(20) NOT NULL DEFAULT 'RUNNING',
        source_file VARCHAR(255) NULL,
        source_sha256 CHAR(64) NULL,
        rows_read INT NOT NULL DEFAULT 0,
        rows_inserted INT NOT NULL DEFAULT 0,
        rows_skipped INT NOT NULL DEFAULT 0,
        funds_matched INT NOT NULL DEFAULT 0,
        funds_ambiguous INT NOT NULL DEFAULT 0,
        funds_unmatched INT NOT NULL DEFAULT 0,
        details_json LONGTEXT NULL,
        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
        KEY idx_batch (batch)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4""",
]


def ensure_tables(conn):
    with conn.cursor() as cur:
        for ddl in DDL:
            cur.execute(ddl)


# ---------------------------------------------------------------------------
# Lecture du classeur
# ---------------------------------------------------------------------------
def read_workbook(xlsx_path):
    """Lit l'onglet Données. Verifie les en-tetes AVANT lecture (anti SCHEMA_DRIFT)."""
    if openpyxl is None:
        raise RuntimeError("openpyxl non installe")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_DATA not in wb.sheetnames:
        raise RuntimeError(f"SCHEMA_DRIFT: onglet '{SHEET_DATA}' absent ({wb.sheetnames})")
    ws = wb[SHEET_DATA]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    missing = [c for c in EXPECTED_COLUMNS if c not in header]
    if missing:
        raise RuntimeError(f"SCHEMA_DRIFT: colonnes absentes {missing}")
    idx = {c: header.index(c) for c in EXPECTED_COLUMNS}

    rows = []
    for raw in it:
        g = lambda c: raw[idx[c]] if idx[c] < len(raw) else None
        vdate = to_iso(g("Date de valorisation"))
        name = g("Fonds")
        if not vdate or not name or not str(name).strip():
            continue  # ligne inexploitable : ignoree, jamais reconstituee
        name = str(name).strip()
        if normalize_name(name) in ("TOTAL", "SUB TOTAL", "SUBTOTAL", "GRAND TOTAL"):
            continue  # faux fonds
        rows.append({
            "valuation_date": vdate,
            "fund_name_raw": name,
            "fund_key": normalize_name(name),
            "manager_raw": (str(g("Gestionnaire")).strip() if g("Gestionnaire") else None),
            "category_sec": (str(g("Catégorie SEC")).strip() if g("Catégorie SEC") else None),
            "net_assets_ngn": to_num(g("Actif net total (NGN)")),
            "net_assets_usd": to_num(g("Actif net total (USD)")),
            "unit_price_ngn": to_num(g("VL / Unit Price (NGN)")),
            "bid_price_ngn": to_num(g("Bid Price (NGN)")),
            "bid_price_usd": to_num(g("Bid Price (USD)")),
            "offer_price_ngn": to_num(g("Offer Price (NGN)")),
            "offer_price_usd": to_num(g("Offer Price (USD)")),
            "quality_status": (str(g("Statut qualité")).strip() if g("Statut qualité") else "OK"),
            "quality_note": (str(g("Note qualité")).strip() if g("Note qualité") else None),
            "has_conflict": 1 if str(g("Conflit source")).strip().lower() in ("oui", "yes", "1", "true") else 0,
            "report_date": to_iso(g("Date du rapport source")),
            "sec_document_id": (str(g("Document SEC ID")).strip() if g("Document SEC ID") else None),
            "source_file": (str(g("Fichier source")).strip() if g("Fichier source") else None),
            "source_url": (str(g("URL source")).strip() if g("URL source") else None),
        })
    for r in rows:
        r["row_hash"] = hashlib.sha1(
            f"{r['valuation_date']}|{r['fund_key']}|{r['sec_document_id']}".encode()
        ).hexdigest()
    return rows


# ---------------------------------------------------------------------------
# Resolution d'identite — jamais de fusion automatique en cas d'ambiguite
# ---------------------------------------------------------------------------
def load_nigeria_funds(conn):
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, nom_fond, societe_gestion FROM fond_investissements WHERE pays=%s",
            (PAYS,))
        funds = cur.fetchall()
    by_norm, by_compact = {}, {}
    for f in funds:
        by_norm.setdefault(normalize_name(f["nom_fond"]), f)
        by_compact.setdefault(compact_key(f["nom_fond"]), f)
    return funds, by_norm, by_compact


def match_fund(fund_key, name_raw, funds, by_norm, by_compact):
    """exact -> compact -> fuzzy (>=93 ET ecart >=3). Sinon AMBIGUOUS/UNMATCHED.
    Ne fusionne jamais automatiquement une correspondance douteuse."""
    if fund_key in by_norm:
        return by_norm[fund_key]["id"], "MATCHED_EXACT", 100.0, None
    ck = compact_key(name_raw)
    if ck in by_compact:
        return by_compact[ck]["id"], "MATCHED_COMPACT", 99.0, None
    if fuzz is not None and funds:
        best_id, best, second = None, 0.0, 0.0
        for f in funds:
            sc = fuzz.token_sort_ratio(fund_key, normalize_name(f["nom_fond"]))
            if sc > best:
                second, best, best_id = best, sc, f["id"]
            elif sc > second:
                second = sc
        if best >= 93 and (best - second) >= 3:
            return best_id, "MATCHED_FUZZY", best, None
        if best >= 85:
            return None, "AMBIGUOUS", best, f"score {best:.0f} trop proche du 2e ({second:.0f})"
    return None, "UNMATCHED", None, "aucun candidat au-dessus du seuil"


# ---------------------------------------------------------------------------
# Comparaison staging vs production — classification, aucune ecriture
# ---------------------------------------------------------------------------
def classify(obs, prod_row):
    """Compare une observation officielle a la ligne de production correspondante.

    Ne decide JAMAIS qu'une valeur doit etre remplacee : produit seulement un
    classement auditable. `value` en production est de nature ambigue (VL, Bid
    ou Offer selon l'epoque) : on le compare donc a TOUTES les mesures publiees
    pour la date, et on signale a quelle mesure il correspond reellement.
    """
    if prod_row is None:
        return "ABSENT_EN_PROD", None
    pv = prod_row.get("value")
    if pv is None:
        return "PROD_SANS_VALEUR", None
    candidates = {
        "unit_price_ngn": obs["unit_price_ngn"], "bid_price_ngn": obs["bid_price_ngn"],
        "offer_price_ngn": obs["offer_price_ngn"], "bid_price_usd": obs["bid_price_usd"],
        "offer_price_usd": obs["offer_price_usd"],
    }
    for label, v in candidates.items():
        if v is not None and abs(float(pv) - float(v)) < 0.0001:
            return ("IDENTIQUE" if label == "unit_price_ngn" else "MESURE_DIFFERENTE"), label
    return "ECART_VALEUR", None


def shift_analysis(conn, rows, sample_limit=25):
    """Teste l'hypothese du decalage d'une periode, SANS RIEN MODIFIER.

    Pour chaque ligne de production (fonds, date D) on cherche si `value`
    correspond a une mesure publiee :
      - a la date D              -> MATCH_DATE_COURANTE (pas de decalage)
      - a la date SEC precedente -> MATCH_DATE_PRECEDENTE (decalage confirme)
      - nulle part               -> AUCUNE_CORRESPONDANCE
    Le rapport est ventile par annee, categorie et devise afin de satisfaire
    l'exigence de demonstration sur plusieurs annees/structures/categories
    avant toute correction de masse. AUCUNE ecriture.
    """
    by_fund = {}
    for r in rows:
        if r.get("matched_fund_id"):
            by_fund.setdefault(r["matched_fund_id"], {})[r["valuation_date"]] = r

    def measures(o):
        return {"unit_price_ngn": o["unit_price_ngn"], "bid_price_ngn": o["bid_price_ngn"],
                "offer_price_ngn": o["offer_price_ngn"], "bid_price_usd": o["bid_price_usd"],
                "offer_price_usd": o["offer_price_usd"]}

    def find(o, val):
        if o is None:
            return None
        for lbl, v in measures(o).items():
            if v is not None and abs(float(val) - float(v)) < 0.0001:
                return lbl
        return None

    stats, by_year, by_cat, by_cur, samples = {}, {}, {}, {}, []
    with conn.cursor() as cur:
        for fund_id, obs_by_date in by_fund.items():
            dates_sorted = sorted(obs_by_date)
            cur.execute("SELECT date, value FROM valorisations WHERE fund_id=%s AND value IS NOT NULL",
                        (fund_id,))
            for prod in cur.fetchall():
                pdate = prod["date"].strftime("%Y-%m-%d") if hasattr(prod["date"], "strftime") else str(prod["date"])
                pval = prod["value"]
                cur_obs = obs_by_date.get(pdate)
                prev_date = None
                for dd in reversed(dates_sorted):          # derniere date SEC < pdate
                    if dd < pdate:
                        prev_date = dd
                        break
                prev_obs = obs_by_date.get(prev_date) if prev_date else None

                hit_cur, hit_prev = find(cur_obs, pval), find(prev_obs, pval)
                if hit_cur:
                    verdict, lbl, src = "MATCH_DATE_COURANTE", hit_cur, pdate
                elif hit_prev:
                    verdict, lbl, src = "MATCH_DATE_PRECEDENTE", hit_prev, prev_date
                else:
                    verdict, lbl, src = "AUCUNE_CORRESPONDANCE", None, None

                stats[verdict] = stats.get(verdict, 0) + 1
                ref = cur_obs or prev_obs
                yr = pdate[:4]
                by_year.setdefault(yr, {})[verdict] = by_year.setdefault(yr, {}).get(verdict, 0) + 1
                if ref:
                    cat = (ref.get("category_sec") or "?")[:38]
                    by_cat.setdefault(cat, {})[verdict] = by_cat.setdefault(cat, {}).get(verdict, 0) + 1
                if lbl:
                    cur_code = "USD" if lbl.endswith("_usd") else "NGN"
                    by_cur.setdefault(cur_code, {})[verdict] = by_cur.setdefault(cur_code, {}).get(verdict, 0) + 1
                if verdict == "MATCH_DATE_PRECEDENTE" and len(samples) < sample_limit and ref:
                    samples.append({"fund_id": fund_id, "fonds": ref["fund_name_raw"],
                                    "date_en_base": pdate, "valeur_en_base": float(pval),
                                    "correspond_a": lbl, "date_sec_reelle": src,
                                    "categorie": ref.get("category_sec")})
    return stats, by_year, by_cat, by_cur, samples


def build_report(conn, rows, batch):
    """Rapport de comparaison ligne a ligne. LECTURE SEULE sur valorisations."""
    stats = {"total": 0, "sans_fonds": 0}
    detail = {}
    with conn.cursor() as cur:
        for r in rows:
            if not r.get("matched_fund_id"):
                stats["sans_fonds"] += 1
                continue
            cur.execute(
                "SELECT value, actif_net FROM valorisations WHERE fund_id=%s AND date=%s LIMIT 1",
                (r["matched_fund_id"], r["valuation_date"]))
            status, measure = classify(r, cur.fetchone())
            r["compare_status"] = status
            stats["total"] += 1
            stats[status] = stats.get(status, 0) + 1
            if measure:
                detail[measure] = detail.get(measure, 0) + 1
    return stats, detail


# ---------------------------------------------------------------------------
# Selftest
# ---------------------------------------------------------------------------
def selftest():
    assert to_num("1,234,567.89") == 1234567.89
    assert to_num("1 234 567.89") == 1234567.89
    assert to_num("1.234.567,89") == 1234567.89
    assert to_num("1'234'567.89") == 1234567.89
    assert to_num("(1,234.56)") == -1234.56
    assert to_num("₦1,234.56") == 1234.56
    assert to_num("USD 1,234.56") == 1234.56
    assert to_num("1.23E+09") == 1.23e9
    assert to_num("1234567,89") == 1234567.89
    assert to_num("0") == 0.0 and to_num("0,00") == 0.0      # zero publie preserve
    for v in ("N/A", "NA", "NIL", "NONE", "ND", "-", "--", "", None):
        assert to_num(v) is None, v
    assert to_num("1 234,56") == 1234.56                 # espace insecable
    assert to_num("1 234.56") == 1234.56                 # espace fine
    assert to_num(118.9768) == 118.9768                       # natif Excel intact
    assert normalize_name("ACAP Canary Growth Fund") == "ACAP CANARY GROWTH FUND"
    assert compact_key("GDL CanaryGrowth Fund") == compact_key("GDL Canary Growth Fund")
    assert normalize_name("A & B Fund") == "A AND B FUND"
    assert to_iso(datetime(2026, 7, 10)) == "2026-07-10"
    assert to_iso(None) is None
    # classify : une valeur de prod egale a l'Offer publie n'est PAS une VL
    obs = {"unit_price_ngn": None, "bid_price_ngn": 189.9601, "offer_price_ngn": 195.5378,
           "bid_price_usd": None, "offer_price_usd": None}
    assert classify(obs, {"value": 195.5378}) == ("MESURE_DIFFERENTE", "offer_price_ngn")
    assert classify(obs, None)[0] == "ABSENT_EN_PROD"
    obs2 = {"unit_price_ngn": 111.07, "bid_price_ngn": None, "offer_price_ngn": None,
            "bid_price_usd": None, "offer_price_usd": None}
    assert classify(obs2, {"value": 111.07}) == ("IDENTIQUE", "unit_price_ngn")
    print("SELFTEST OK — parsing numerique, normalisation et classification valides")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Chargeur classeur SEC Nigeria 2011-2026")
    ap.add_argument("--xlsx", help="chemin du classeur .xlsx")
    ap.add_argument("--dry-run", action="store_true", help="simulation (defaut)")
    ap.add_argument("--execute", action="store_true", help="ecrit dans les tables sec_ng_* (STAGING uniquement)")
    ap.add_argument("--report", action="store_true", help="rapport de comparaison avec la production (lecture seule)")
    ap.add_argument("--shift-analysis", action="store_true", help="teste l'hypothese du decalage de date (lecture seule)")
    ap.add_argument("--selftest", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    if args.selftest:
        selftest()
        return 0
    if not args.xlsx:
        ap.error("--xlsx requis (ou --selftest)")
    if not args.execute:
        args.dry_run = True
        log.info("MODE DRY-RUN — aucune ecriture (meme en staging)")

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        log.error("Fichier introuvable : %s", xlsx)
        return 1
    sha = hashlib.sha256(xlsx.read_bytes()).hexdigest()
    batch = f"SECNG_{datetime.now():%Y%m%d_%H%M%S}"
    log.info("Classeur : %s (sha256 %s...)", xlsx.name, sha[:16])

    rows = read_workbook(xlsx)
    log.info("Lignes exploitables lues : %d", len(rows))
    dates = sorted({r["valuation_date"] for r in rows})
    log.info("Periode : %s -> %s (%d dates distinctes, %d cles de fonds)",
             dates[0], dates[-1], len(dates), len({r["fund_key"] for r in rows}))

    conn = db_connect()
    try:
        funds, by_norm, by_compact = load_nigeria_funds(conn)
        log.info("Referentiel Nigeria en base : %d fonds", len(funds))

        cache, m_stats = {}, {}
        for r in rows:
            key = r["fund_key"]
            if key not in cache:
                cache[key] = match_fund(key, r["fund_name_raw"], funds, by_norm, by_compact)
            fid, status, conf, reason = cache[key]
            r["matched_fund_id"], r["match_status"] = fid, status
            r["match_confidence"], r["review_reason"] = conf, reason
            m_stats[status] = m_stats.get(status, 0) + 1

        log.info("--- Resolution d'identite (par observation) ---")
        for k, v in sorted(m_stats.items(), key=lambda x: -x[1]):
            log.info("  %-16s : %d", k, v)
        keys_by_status = {}
        for key, (fid, status, conf, reason) in cache.items():
            keys_by_status.setdefault(status, []).append(key)
        log.info("--- Resolution d'identite (par cle de fonds, %d cles) ---", len(cache))
        for k, v in sorted(keys_by_status.items(), key=lambda x: -len(x[1])):
            log.info("  %-16s : %d cles", k, len(v))
        for st in ("AMBIGUOUS", "UNMATCHED"):
            for key in sorted(keys_by_status.get(st, []))[:15]:
                log.info("    [%s] %s", st, key)

        if args.report:
            log.info("--- Comparaison avec la production (LECTURE SEULE) ---")
            stats, detail = build_report(conn, rows, batch)
            for k, v in sorted(stats.items(), key=lambda x: -x[1] if isinstance(x[1], int) else 0):
                log.info("  %-20s : %s", k, v)
            if detail:
                log.info("  Mesure reellement stockee dans valorisations.value :")
                for k, v in sorted(detail.items(), key=lambda x: -x[1]):
                    log.info("    %-18s : %d", k, v)

        shift_out = None
        if args.shift_analysis:
            log.info("--- Analyse du decalage de date (LECTURE SEULE) ---")
            st, by_year, by_cat, by_cur, samples = shift_analysis(conn, rows)
            tot = sum(st.values()) or 1
            for k, v in sorted(st.items(), key=lambda x: -x[1]):
                log.info("  %-24s : %6d  (%5.1f%%)", k, v, 100.0*v/tot)
            log.info("  --- par annee ---")
            for yr in sorted(by_year):
                d = by_year[yr]
                log.info("    %s : courante=%-5d precedente=%-5d aucune=%-5d", yr,
                         d.get("MATCH_DATE_COURANTE",0), d.get("MATCH_DATE_PRECEDENTE",0),
                         d.get("AUCUNE_CORRESPONDANCE",0))
            log.info("  --- par categorie SEC (top 10) ---")
            for cat, d in sorted(by_cat.items(), key=lambda x: -sum(x[1].values()))[:10]:
                log.info("    %-40s courante=%-5d precedente=%-5d aucune=%-5d", cat,
                         d.get("MATCH_DATE_COURANTE",0), d.get("MATCH_DATE_PRECEDENTE",0),
                         d.get("AUCUNE_CORRESPONDANCE",0))
            log.info("  --- par devise ---")
            for c, d in sorted(by_cur.items()):
                log.info("    %-4s courante=%-5d precedente=%-5d", c,
                         d.get("MATCH_DATE_COURANTE",0), d.get("MATCH_DATE_PRECEDENTE",0))
            log.info("  --- echantillons de decalage confirme ---")
            for smp in samples[:12]:
                log.info("    [%s] %s : base %s=%s -> reellement %s du %s",
                         smp["fund_id"], smp["fonds"][:34], smp["date_en_base"],
                         smp["valeur_en_base"], smp["correspond_a"], smp["date_sec_reelle"])
            shift_out = {"stats": st, "par_annee": by_year, "par_categorie": by_cat,
                         "par_devise": by_cur, "echantillons": samples}

        if args.execute:
            ensure_tables(conn)
            ins = skip = 0
            with conn.cursor() as cur:
                for r in rows:
                    try:
                        cur.execute("""INSERT INTO sec_ng_observations
                            (valuation_date, fund_name_raw, fund_key, manager_raw, category_sec,
                             net_assets_ngn, net_assets_usd, unit_price_ngn, bid_price_ngn,
                             bid_price_usd, offer_price_ngn, offer_price_usd, quality_status,
                             quality_note, has_conflict, report_date, sec_document_id,
                             source_file, source_url, row_hash, matched_fund_id, match_status,
                             match_confidence, compare_status, load_batch)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (r["valuation_date"], r["fund_name_raw"], r["fund_key"], r["manager_raw"],
                             r["category_sec"], r["net_assets_ngn"], r["net_assets_usd"],
                             r["unit_price_ngn"], r["bid_price_ngn"], r["bid_price_usd"],
                             r["offer_price_ngn"], r["offer_price_usd"], r["quality_status"],
                             r["quality_note"], r["has_conflict"], r["report_date"],
                             r["sec_document_id"], r["source_file"], r["source_url"], r["row_hash"],
                             r.get("matched_fund_id"), r.get("match_status"),
                             r.get("match_confidence"), r.get("compare_status"), batch))
                        ins += 1
                    except pymysql.err.IntegrityError:
                        skip += 1                      # deja en staging : idempotent
                for key, (fid, status, conf, reason) in cache.items():
                    sub = [r for r in rows if r["fund_key"] == key]
                    cur.execute("""INSERT INTO sec_ng_fund_aliases
                        (fund_key, fund_name_raw, normalized_name, compact_name, manager_raw,
                         fund_id, match_status, confidence, review_reason, first_seen, last_seen, observations)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON DUPLICATE KEY UPDATE last_seen=VALUES(last_seen),
                          observations=VALUES(observations), updated_at=NOW()""",
                        (key, sub[-1]["fund_name_raw"], key, compact_key(sub[-1]["fund_name_raw"]),
                         sub[-1]["manager_raw"], fid, status, conf, reason,
                         min(r["valuation_date"] for r in sub),
                         max(r["valuation_date"] for r in sub), len(sub)))
                cur.execute("""INSERT INTO sec_ng_load_logs
                    (batch, started_at, finished_at, status, source_file, source_sha256,
                     rows_read, rows_inserted, rows_skipped, funds_matched, funds_ambiguous,
                     funds_unmatched, details_json)
                    VALUES (%s,NOW(),NOW(),'SUCCESS',%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (batch, xlsx.name, sha, len(rows), ins, skip,
                     len(keys_by_status.get("MATCHED_EXACT", [])) + len(keys_by_status.get("MATCHED_COMPACT", [])) + len(keys_by_status.get("MATCHED_FUZZY", [])),
                     len(keys_by_status.get("AMBIGUOUS", [])),
                     len(keys_by_status.get("UNMATCHED", [])),
                     json.dumps({"match_stats": m_stats}, default=str)[:60000]))
            log.info("STAGING : %d inserees, %d deja presentes (idempotent)", ins, skip)
        else:
            log.info(">>> DRY-RUN : aucune ecriture. Utiliser --execute pour remplir le staging. <<<")

        REPORT_DIR.mkdir(parents=True, exist_ok=True)
        rp = REPORT_DIR / f"sec_ng_load_{batch}.json"
        rp.write_text(json.dumps({
            "batch": batch, "source": xlsx.name, "sha256": sha,
            "rows": len(rows), "dates": len(dates),
            "periode": [dates[0], dates[-1]],
            "match_stats_observations": m_stats,
            "match_stats_cles": {k: len(v) for k, v in keys_by_status.items()},
            "cles_ambigues": sorted(keys_by_status.get("AMBIGUOUS", [])),
            "cles_non_resolues": sorted(keys_by_status.get("UNMATCHED", [])),
            "analyse_decalage": shift_out,
        }, indent=2, ensure_ascii=False, default=str))
        log.info("Rapport : %s", rp)
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
