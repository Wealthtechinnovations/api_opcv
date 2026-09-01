#!/usr/bin/env python3
"""
La SEC Nigeria publie-t-elle une colonne en dollars, et depuis quand ?

POURQUOI CE SCRIPT
------------------
La devise emise par l extracteur passe de 0 % de dollars en 2022, 2023, 2024 et
2025 a 51 % en 2026, sans aucun degrade. Deux causes possibles, qui n appellent
pas le meme remede :

  - la SEC a commence a publier une colonne dollar en 2026 : alors l historique
    ancien n en contient pas, et rien ne peut le rendre dollar sans fabriquer
    des valeurs que personne n a publiees ;
  - ou l extracteur ne sait lire que le format 2026 : alors la donnee dollar
    existe dans les fichiers anciens, et amputer la serie detruirait du
    recuperable.

Trois fichiers regardes a la main donnaient tous des en-tetes « (N) », y compris
celui de 2026 — mais c etait le fichier du 10 avril, deja connu pour n avoir
aucune colonne dollar. L echantillon designait precisement le cas particulier.
Conclure la-dessus serait refaire l erreur que ce chantier paie depuis des mois.

Ce script ouvre TOUS les fichiers et compte, par annee, combien portent au moins
un en-tete en dollars. Il ne juge pas les valeurs : seulement la presence d une
colonne. C est la source qui repond.

COUVERTURE — corrigee le 2026-08-29. La premiere version ne balayait que
`*/*.xlsx` et a lu 340 fichiers. Or `PROMPT_NIGERIA_ZERO_REGRESSION_V2_2.md`
(ligne 61) recense **686 fichiers officiels SEC sur les pages annuelles
2011-2026**, et le cache du serveur contient aussi des `.xls` anciens
qu openpyxl ne sait pas ouvrir. Conclure « aucun dollar avant 2026 » sur la
moitie du corpus, en ayant ecarte precisement les fichiers les plus anciens,
n etait pas une mesure : c etait un echantillon presente comme un recensement.

Le balayage couvre desormais toutes les extensions et toute profondeur, et
affiche explicitement ce qu il n a PAS pu lire. Un fichier illisible doit
apparaitre comme tel, jamais disparaitre du denominateur.

LECTURE SEULE : aucun fichier n est modifie, aucune base n est touchee.

USAGE
  python3 scripts/diag/scan_sec_headers.py [repertoire]     # defaut : sec_ng_downloads
"""

import collections
import glob
import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl indisponible — ce balayage ne peut pas s executer.")
    sys.exit(0)

# « ($) », « (USD) », « (US$) ». Le symbole seul ne suffit pas : un montant
# « $1,234 » dans une cellule de texte n est pas un en-tete de colonne. La
# parenthese est ce qui distingue l en-tete de la donnee, comme dans « (N) ».
MARQUEUR_USD = re.compile(r"\(\s*(?:\$|USD|US\$)\s*\)", re.I)
MARQUEUR_NGN = re.compile(r"\(\s*(?:N|NGN|NAIRA)\s*\)", re.I)

# Les en-tetes vivent dans les toutes premieres lignes ; lire au-dela couterait
# du temps sans rien apprendre.
LIGNES_ENTETE = 4


def entetes_du_fichier(chemin):
    """Les cellules non vides des premieres lignes de chaque feuille."""
    cellules = []
    wb = load_workbook(chemin, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            for ligne in ws.iter_rows(min_row=1, max_row=LIGNES_ENTETE, values_only=True):
                for cel in ligne:
                    if cel is not None and str(cel).strip():
                        cellules.append(str(cel))
    finally:
        wb.close()
    return cellules


def main():
    racine = sys.argv[1] if len(sys.argv) > 1 else "sec_ng_downloads"
    # Toute profondeur, toutes extensions de classeur. `*/*.xlsx` ratait a la
    # fois les .xls anciens et tout fichier range autrement.
    tous = sorted(
        c for c in glob.glob(os.path.join(racine, "**", "*"), recursive=True)
        if os.path.isfile(c) and os.path.splitext(c)[1].lower() in (".xlsx", ".xlsm", ".xls")
    )
    fichiers = [c for c in tous if os.path.splitext(c)[1].lower() != ".xls"]
    anciens_xls = [c for c in tous if os.path.splitext(c)[1].lower() == ".xls"]

    if not tous:
        print(f"Aucun classeur sous {racine}")
        return

    print(f"\n=== EN-TETES SEC — {len(tous)} classeurs sous {racine} ===\n")
    print(f"  {len(fichiers)} lisibles directement (.xlsx / .xlsm)")
    print(f"  {len(anciens_xls)} au format .xls ancien — openpyxl ne les ouvre pas")
    # Le prompt de reference recense le corpus officiel. L ecart se dit, il ne
    # se tait pas : une couverture partielle presentee comme complete est
    # exactement ce qui a coute des mois a ce chantier.
    ATTENDU = 686
    manquants = ATTENDU - len(tous)
    if manquants > 0:
        print(f"  {manquants} fichier(s) ABSENTS du cache — le prompt V2.2 en recense {ATTENDU} (2011-2026)")
        print("  Toute conclusion ci-dessous ne vaut donc que pour les fichiers PRESENTS.")
    print()

    stats = collections.defaultdict(lambda: {"total": 0, "usd": 0, "ngn": 0, "erreurs": 0})
    exemples_usd = []
    premiers_usd = {}

    # L annee vient du nom du fichier quand il la porte, sinon du repertoire :
    # se fier a la position dans le chemin cassait des que l arborescence
    # changeait de profondeur.
    def annee_de(chemin):
        m = re.search(r"(20\d{2})", os.path.basename(chemin))
        if m:
            return m.group(1)
        for partie in chemin.split(os.sep):
            if re.fullmatch(r"20\d{2}", partie):
                return partie
        return "?"

    for chemin in fichiers:
        annee = annee_de(chemin)
        b = stats[annee]
        b["total"] += 1
        try:
            cellules = entetes_du_fichier(chemin)
        except Exception:
            b["erreurs"] += 1
            continue

        a_usd = False
        a_ngn = False
        for c in cellules:
            if not a_usd and MARQUEUR_USD.search(c):
                a_usd = True
                if len(exemples_usd) < 8:
                    exemples_usd.append(f"{os.path.basename(chemin)} — « {c.strip()[:58]} »")
            if not a_ngn and MARQUEUR_NGN.search(c):
                a_ngn = True
            if a_usd and a_ngn:
                break

        if a_usd:
            b["usd"] += 1
            # Tous les noms, pas le premier rencontre : `sorted(glob)` trie par
            # ordre ALPHABETIQUE, donc « 10th_July » precede « 15th_May ».
            # Annoncer « premier fichier a colonne dollar » d apres cet ordre
            # donnerait une date fausse — et c est la date de bascule qui decide
            # du perimetre applicable.
            premiers_usd.setdefault(annee, []).append(os.path.basename(chemin))
        if a_ngn:
            b["ngn"] += 1

    print(f"  {'annee':6} {'fichiers':>9} {'avec ($)':>9} {'avec (N)':>9} {'illisibles':>11}   part dollar")
    print(f"  {'-'*6} {'-'*9} {'-'*9} {'-'*9} {'-'*11}   -----------")
    for annee in sorted(stats):
        b = stats[annee]
        lisibles = b["total"] - b["erreurs"]
        pct = (b["usd"] / lisibles * 100) if lisibles else 0
        print(f"  {annee:6} {b['total']:9} {b['usd']:9} {b['ngn']:9} {b['erreurs']:11}   {pct:.1f} %")

    total = sum(b["total"] for b in stats.values())
    total_usd = sum(b["usd"] for b in stats.values())
    print(f"\n  Total : {total_usd} fichier(s) sur {total} lus portent au moins un en-tete en dollars.")

    if anciens_xls:
        par_annee_xls = collections.Counter(annee_de(c) for c in anciens_xls)
        print(f"\n  NON LUS — {len(anciens_xls)} fichiers .xls, par annee :")
        for an in sorted(par_annee_xls):
            print(f"    {an} : {par_annee_xls[an]}")
        print("  Ils demandent une conversion LibreOffice, deja utilisee par l extracteur.")
        print("  Tant qu ils ne sont pas lus, aucune affirmation sur ces annees n est fondee.")

    if premiers_usd:
        # Le mois est extrait du nom pour classer chronologiquement ; a defaut on
        # affiche la liste brute plutot qu une date qu on ne sait pas ordonner.
        MOIS = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]
        def rang(nom):
            for i, m in enumerate(MOIS):
                if m.lower() in nom.lower():
                    jour = re.search(r"(\d{1,2})(?:st|nd|rd|th)", nom)
                    return (i, int(jour.group(1)) if jour else 0)
            return (99, 0)
        print("\n  Fichiers a colonne dollar, par annee (ordre chronologique) :")
        for annee in sorted(premiers_usd):
            noms = sorted(premiers_usd[annee], key=rang)
            print(f"    {annee} : {len(noms)} fichier(s), du premier au dernier")
            print(f"      debut : {noms[0]}")
            print(f"      fin   : {noms[-1]}")

    if exemples_usd:
        print("\n  Exemples d en-tetes en dollars :")
        for e in exemples_usd:
            print(f"    {e}")
    else:
        print("\n  AUCUN en-tete en dollars dans aucun fichier.")
        print("  Consequence : l historique ne peut pas devenir dollar sans fabriquer")
        print("  des valeurs. Seule la periode ou la SEC en publie est concernee.")

    print()


if __name__ == "__main__":
    main()
