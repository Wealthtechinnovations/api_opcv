#!/usr/bin/env python3
"""Affiche la structure reelle des en-tetes d un fichier SEC Nigeria.

LECTURE SEULE. N ecrit rien, ne touche ni la base ni les CSV de production.

POURQUOI
--------
Le correctif du lot AF faisait lire la devise dans l en-tete de la colonne de
prix utilisee. La mesure (lot AG) montre qu aucune ligne n y est parvenue :
`vl_currency_source` vaut `inferred_*` partout, jamais `column_header`. Les
en-tetes de ces fichiers ne portent donc PAS le marqueur de devise a l endroit
suppose — l hypothese « Offer Price (USD) » etait fausse pour les fichiers 2026.

Avant de coder une huitieme hypothese, on regarde ou la devise se trouve
reellement : ligne d en-tete groupee au-dessus, cellule fusionnee, ou ailleurs.

USAGE
    python3 scripts/diag/dump_sec_header_structure.py [chemin.xlsx] [nb_lignes]

Sans argument, prend le premier .xlsx du cache `sec_ng_downloads/`.
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl absent — ce diagnostic doit tourner sur le serveur.")
    sys.exit(2)

RACINE = Path(__file__).resolve().parent.parent.parent


def texte(v) -> str:
    if v is None:
        return ""
    return str(v).strip().replace("\n", " ")[:26]


def main() -> int:
    if len(sys.argv) > 1:
        chemin = Path(sys.argv[1])
    else:
        cache = RACINE / "sec_ng_downloads"
        candidats = sorted(cache.glob("**/*.xlsx")) if cache.exists() else []
        # Un fichier recent et volumineux a le plus de chances de porter la
        # structure moderne a blocs larges.
        candidats = [c for c in candidats if "2026" in c.name or "2026" in str(c.parent)]
        if not candidats:
            print("Aucun .xlsx 2026 dans sec_ng_downloads/")
            return 1
        chemin = max(candidats, key=lambda p: p.stat().st_size)

    nb = int(sys.argv[2]) if len(sys.argv) > 2 else 14
    print(f"\nFichier : {chemin.name}")
    print(f"Taille  : {chemin.stat().st_size / 1024:.0f} Ko\n")

    wb = load_workbook(chemin, data_only=True, read_only=True)
    for ws in wb.worksheets:
        lignes = []
        for i, row in enumerate(ws.iter_rows(max_row=nb, values_only=True)):
            lignes.append(list(row))
        if not lignes:
            continue
        largeur = max(len(l) for l in lignes)
        print(f"=== Feuille « {ws.title} » — {nb} premieres lignes, {largeur} colonnes ===\n")

        # Reperer les colonnes qui portent un mot-cle de prix : c est autour
        # d elles que la devise doit se trouver.
        interessantes = set()
        for r, l in enumerate(lignes):
            for c, v in enumerate(l):
                t = texte(v).upper()
                if any(k in t for k in ("PRICE", "NAV", "USD", "NGN", "$", "NAIRA", "DOLLAR")):
                    interessantes.add(c)
        cols = sorted(interessantes)[:22]
        if not cols:
            print("   (aucune colonne portant PRICE / NAV / devise dans cette zone)\n")
            continue

        print("   ligne | " + " | ".join(f"c{c:<3}" for c in cols))
        print("   ------+" + "-+".join("-" * 5 for _ in cols))
        for r, l in enumerate(lignes):
            cellules = []
            for c in cols:
                cellules.append(texte(l[c]) if c < len(l) else "")
            if any(cellules):
                print(f"   {r:5} | " + " | ".join(f"{x[:26]:<26}" for x in cellules))
        print()

        # Les cellules fusionnees portent souvent le libelle de devise, stocke
        # uniquement dans la premiere colonne de la fusion.
        try:
            wb2 = load_workbook(chemin, data_only=True)
            ws2 = wb2[ws.title]
            fusions = [str(m) for m in ws2.merged_cells.ranges][:25]
            print(f"   Cellules fusionnees ({len(fusions)} premieres) : {', '.join(fusions) if fusions else 'aucune'}\n")
        except Exception as e:
            print(f"   (fusions illisibles : {e})\n")
        # --- Lignes d un fonds en devise etrangere ---
        #
        # Les lignes ci-dessus sont des fonds en naira (colonnes ($) a N/A).
        # Pour trancher #73 il faut voir un fonds DOLLAR : quelle colonne porte
        # reellement la valeur que l extracteur retient ?
        print("   === Fonds en devise etrangere : valeur par colonne ===\n")
        entetes = {}
        for r, l in enumerate(lignes[:6]):
            for c, v in enumerate(l):
                t = texte(v)
                if t and any(k in t.upper() for k in ("PRICE", "NAV")):
                    entetes.setdefault(c, t)

        trouves = 0
        for row in ws.iter_rows(max_row=400, values_only=True):
            nom = ""
            for v in list(row)[:6]:
                t = texte(v)
                if t and any(k in t.upper() for k in ("DOLLAR", "EUROBOND")):
                    nom = t
                    break
            if not nom:
                continue
            trouves += 1
            if trouves > 3:
                break
            print(f"   --- {nom} ---")
            for c in sorted(entetes):
                val = texte(row[c]) if c < len(row) else ""
                if val and val.upper() not in ("N/A", "NA", "-"):
                    print(f"      c{c:<3} {entetes[c][:24]:<24} = {val}")
            print()
        if not trouves:
            print("   (aucun fonds dollar/eurobond dans les 400 premieres lignes)\n")

        break  # une feuille suffit a comprendre la structure
    return 0


if __name__ == "__main__":
    sys.exit(main())
